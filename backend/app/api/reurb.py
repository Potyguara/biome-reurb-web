from datetime import datetime, timezone
from uuid import UUID
import mimetypes

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Request,
    UploadFile,
    status,
)
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.access import ProjectUser
from app.models.lot_geometry import LotGeometry
from app.models.project import Project
from app.models.reurb import (
    Document,
    Lot,
    PhysicalRegistration,
    ProjectOrthomosaic,
    Seal,
    SocialRegistration,
)

from app.models.user import User
from app.schemas.reurb import (
    DocumentResponse,
    DocumentValidateRequest,
    DocumentValidationUpdate,
    LotDeleteCheckResponse,
    LotDocumentResponse,
    LotDocumentUploadResponse,
    LotLinkCandidateResponse,
    LotLinkSealRequest,
    LotLinkSealResponse,
    LotResponse,
    LotReviewResponse,
    LotReviewUpdate,
    PhysicalRegistrationResponse,
    PhysicalRegistrationUpdateRequest,
    ProjectDashboardResponse,
    ProjectMapLotResponse,
    ProjectMapPhysicalResponse,
    ProjectMapProjectResponse,
    ProjectMapResponse,
    ProjectMapSealResponse,
    ProjectMapSealWithoutLotResponse,
    ProjectMapSocialResponse,
    ProjectMapSummaryResponse,
    SealDeleteCheckResponse,
    SealDeleteResponse,
    SealResponse,
    SealUpdateRequest,
    SocialRegistrationResponse,
    SocialRegistrationUpdateRequest,
)
from app.services.audit_service import register_audit_log
from app.services.geospatial_import_service import (
    import_lot_geometries_to_project,
    save_geospatial_upload,
)
from pydantic import BaseModel
import shutil
import uuid
from pathlib import Path

from fastapi.responses import FileResponse
from app.schemas.reurb import SealUpdateRequest

from sqlalchemy import or_
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import csv
import json
import zipfile
from io import StringIO

import html
import tempfile
import shapefile
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
import subprocess

router = APIRouter(prefix="/projects/{project_id}", tags=["Consulta REURB"])
EXPORTS_DIR = Path("storage/exports")
STORAGE_DIR = Path("storage")
ORTHOMOSAICS_DIR = Path("storage/orthomosaics")


def _ensure_gdal_available() -> None:
    if shutil.which("gdalinfo") is None or shutil.which("gdal_translate") is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=(
                "GDAL não encontrado no servidor. Instale com: brew install gdal. "
                "Depois confirme com gdalinfo --version e gdal_translate --version."
            ),
        )


def _run_command(command: list[str], *, user_error_status: int = 500) -> str:
    result = subprocess.run(
        command,
        capture_output=True,
        text=True,
        check=False,
    )

    if result.returncode != 0:
        detail = result.stderr or result.stdout or "Erro ao executar comando GDAL."

        raise HTTPException(
            status_code=user_error_status,
            detail=(
                "Falha ao processar arquivo geoespacial com GDAL.\n\n"
                f"Comando: {' '.join(command)}\n\n"
                f"Detalhe técnico:\n{detail}"
            ),
        )

    return result.stdout


def _extract_crs_label_from_gdal_info(info: dict) -> str | None:
    coordinate_system = info.get("coordinateSystem")

    if not isinstance(coordinate_system, dict):
        return None

    wkt = coordinate_system.get("wkt")

    if isinstance(wkt, str):
        match = re.search(r'ID\["EPSG",\s*(\d+)\]', wkt)

        if match:
            return f"EPSG:{match.group(1)}"

        if "SIRGAS" in wkt.upper():
            return "SIRGAS 2000"

        if "WGS 84" in wkt.upper() or "WGS_1984" in wkt.upper():
            return "WGS 84"

    projjson = coordinate_system.get("projjson")

    if isinstance(projjson, dict):
        id_data = projjson.get("id")

        if isinstance(id_data, dict):
            authority = id_data.get("authority")
            code = id_data.get("code")

            if authority and code:
                return f"{authority}:{code}"

    return "CRS reconhecido pelo GDAL"


def _generate_orthomosaic_preview_with_gdal(
    *,
    source_path: Path,
    preview_path: Path,
    max_size: int = 2500,
) -> dict:
    _ensure_gdal_available()

    info_text = _run_command(
        [
            "gdalinfo",
            "-json",
            str(source_path),
        ],
        user_error_status=status.HTTP_400_BAD_REQUEST,
    )

    try:
        info = json.loads(info_text)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "O arquivo foi recebido, mas o GDAL não conseguiu ler os "
                "metadados como GeoTIFF válido."
            ),
        )

    driver = info.get("driverShortName")

    if driver not in {"GTiff", "COG"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "O arquivo enviado não foi reconhecido como GeoTIFF/COG. "
                "Envie um ortomosaico .tif, .tiff ou .geotiff georreferenciado."
            ),
        )

    coordinate_system = info.get("coordinateSystem")

    if not coordinate_system:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "O ortomosaico não possui sistema de referência definido. "
                "Exporte novamente o GeoTIFF com CRS válido."
            ),
        )

    crs = _extract_crs_label_from_gdal_info(info)

    size = info.get("size") or [None, None]

    try:
        width = int(size[0]) if size and size[0] else None
        height = int(size[1]) if size and size[1] else None
    except Exception:
        width = None
        height = None

    if not width or not height:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Não foi possível identificar largura e altura do ortomosaico.",
        )

    wgs84_extent = info.get("wgs84Extent")
    coordinates = None

    if isinstance(wgs84_extent, dict):
        coordinates = wgs84_extent.get("coordinates")

    lon_values: list[float] = []
    lat_values: list[float] = []

    if isinstance(coordinates, list):
        try:
            ring = coordinates[0]

            for coord in ring:
                lon_values.append(float(coord[0]))
                lat_values.append(float(coord[1]))
        except Exception:
            lon_values = []
            lat_values = []

    if not lon_values or not lat_values:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "O GDAL leu o arquivo, mas não conseguiu transformar os limites "
                "do ortomosaico para latitude/longitude. Verifique se o CRS do "
                "GeoTIFF é válido e reconhecido pelo GDAL."
            ),
        )

    min_lon = min(lon_values)
    max_lon = max(lon_values)
    min_lat = min(lat_values)
    max_lat = max(lat_values)

    if not (-180 <= min_lon <= 180 and -180 <= max_lon <= 180):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Os limites de longitude do ortomosaico ficaram fora do intervalo "
                "válido. Verifique se o sistema de referência do GeoTIFF está correto."
            ),
        )

    if not (-90 <= min_lat <= 90 and -90 <= max_lat <= 90):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Os limites de latitude do ortomosaico ficaram fora do intervalo "
                "válido. Verifique se o sistema de referência do GeoTIFF está correto."
            ),
        )

    preview_path.parent.mkdir(parents=True, exist_ok=True)

    band_count = len(info.get("bands") or [])

    temp_preview_tif = preview_path.with_suffix(".preview_tmp.tif")

    def cleanup_temp() -> None:
        for path in [
            temp_preview_tif,
            Path(str(temp_preview_tif) + ".aux.xml"),
            Path(str(preview_path) + ".aux.xml"),
        ]:
            if path.exists():
                path.unlink(missing_ok=True)

    cleanup_temp()

    rgb_command = [
        "gdal_translate",
        "-of",
        "GTiff",
        "-ot",
        "Byte",
        "-scale",
        "-outsize",
        str(max_size),
        "0",
    ]

    if band_count >= 3:
        rgb_command.extend(["-b", "1", "-b", "2", "-b", "3"])

    rgb_command.extend(
        [
            str(source_path),
            str(temp_preview_tif),
        ]
    )

    grayscale_command = [
        "gdal_translate",
        "-of",
        "GTiff",
        "-ot",
        "Byte",
        "-scale",
        "-outsize",
        str(max_size),
        "0",
        "-b",
        "1",
        str(source_path),
        str(temp_preview_tif),
    ]

    last_error = None

    try:
        _run_command(
            rgb_command,
            user_error_status=status.HTTP_400_BAD_REQUEST,
        )
    except HTTPException as exc:
        last_error = exc

        cleanup_temp()

        try:
            _run_command(
                grayscale_command,
                user_error_status=status.HTTP_400_BAD_REQUEST,
            )
        except HTTPException as fallback_exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "O ortomosaico foi lido, mas não foi possível gerar a prévia "
                    "raster para exibição no mapa. Tente exportar o GeoTIFF como RGB "
                    "8 bits ou COG RGB.\n\n"
                    f"Erro RGB: {last_error.detail}\n\n"
                    f"Erro banda única: {fallback_exc.detail}"
                ),
            )

    if not temp_preview_tif.exists():
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="A prévia temporária do ortomosaico não foi gerada.",
        )

    jpeg_command = [
        "gdal_translate",
        "-of",
        "JPEG",
        "-co",
        "QUALITY=85",
        str(temp_preview_tif),
        str(preview_path),
    ]

    _run_command(
        jpeg_command,
        user_error_status=status.HTTP_400_BAD_REQUEST,
    )

    cleanup_temp()

    if not preview_path.exists():
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="A prévia JPG do ortomosaico não foi gerada.",
        )

    return {
        "crs": str(crs) if crs else None,
        "min_lon": float(min_lon),
        "min_lat": float(min_lat),
        "max_lon": float(max_lon),
        "max_lat": float(max_lat),
        "width": width,
        "height": height,
    }


def _orthomosaic_to_response(item: ProjectOrthomosaic) -> dict:
    return {
        "id": str(item.id),
        "project_id": str(item.project_id),
        "original_filename": item.original_filename,
        "stored_filename": item.stored_filename,
        "crs": item.crs,
        "min_lon": item.min_lon,
        "min_lat": item.min_lat,
        "max_lon": item.max_lon,
        "max_lat": item.max_lat,
        "width": item.width,
        "height": item.height,
        "is_active": item.is_active,
        "created_at": item.created_at.isoformat() if item.created_at else None,
    }


def _resolve_export_document_path(document: Document) -> Path | None:
    raw_path = (document.file_path or "").strip()

    if not raw_path:
        return None

    candidates: list[Path] = []

    direct = Path(raw_path)

    if direct.is_absolute():
        candidates.append(direct)
    else:
        candidates.append(Path(raw_path))
        candidates.append(Path("storage") / raw_path)

    filename = Path(raw_path.replace("\\", "/")).name

    if filename:
        candidates.append(
            Path("storage/documents") / str(document.project_id) / filename
        )

        imports_dir = Path("storage/imports")
        if imports_dir.exists():
            candidates.extend(imports_dir.rglob(filename))

    for candidate in candidates:
        try:
            if candidate.exists() and candidate.is_file():
                return candidate
        except Exception:
            continue

    return None


def _generate_lot_dossier_pdf(
    *,
    project: Project,
    lot: Lot,
    seal: Seal | None,
    social: SocialRegistration | None,
    physical: PhysicalRegistration | None,
    documents_count: int,
    validated_documents_count: int,
) -> bytes:
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    story = []

    title = styles["Title"]
    heading = styles["Heading2"]
    normal = styles["BodyText"]

    story.append(Paragraph("FICHA CADASTRAL INDIVIDUAL DO LOTE - REURB", title))
    story.append(Spacer(1, 0.3 * cm))

    story.append(
        Paragraph(
            f"<b>Projeto:</b> {_safe_export_text(project.name)}<br/>"
            f"<b>Município/UF:</b> {_safe_export_text(project.municipality)}/{_safe_export_text(project.state)}<br/>"
            f"<b>Bairro/Núcleo:</b> {_safe_export_text(project.neighborhood)}",
            normal,
        )
    )

    story.append(Spacer(1, 0.5 * cm))

    def add_section(title_text: str, rows: list[list[str]]) -> None:
        story.append(Paragraph(title_text, heading))

        table = Table(rows, colWidths=[6 * cm, 10 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#166534")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F1F5F9")),
                    ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )

        story.append(table)
        story.append(Spacer(1, 0.4 * cm))

    add_section(
        "1. Identificação do lote",
        [
            ["Campo", "Informação"],
            ["Código do lote", _safe_export_text(lot.code)],
            ["Quadra", _safe_export_text(lot.block, "-")],
            ["Área vetorizada", f"{lot.area_m2 or 0} m²"],
            ["Perímetro", f"{lot.perimeter_m or 0} m"],
            [
                "Status do lote",
                _safe_export_text(getattr(lot, "technical_status", None) or lot.status),
            ],
            [
                "Apto para peças técnicas",
                _yes_no(getattr(lot, "is_ready_for_technical_documents", False)),
            ],
            ["Fonte da geometria", _safe_export_text(lot.source_file, "-")],
        ],
    )

    add_section(
        "2. Selagem vinculada",
        [
            ["Campo", "Informação"],
            [
                "Código da selagem",
                _safe_export_text(seal.seal_code if seal else None, "-"),
            ],
            ["Situação", _safe_export_text(seal.situation if seal else None, "-")],
            [
                "Vínculo geográfico",
                _safe_export_text(seal.geo_link_status if seal else None, "-"),
            ],
            ["Necessita RTK", _yes_no(seal.needs_rtk_validation) if seal else "-"],
            ["Latitude", _safe_export_text(seal.latitude if seal else None, "-")],
            ["Longitude", _safe_export_text(seal.longitude if seal else None, "-")],
        ],
    )

    add_section(
        "3. Cadastro social",
        [
            ["Campo", "Informação"],
            [
                "Responsável",
                _safe_export_text(social.responsible_name if social else None, "-"),
            ],
            ["CPF", _safe_export_text(social.responsible_cpf if social else None, "-")],
            ["RG", _safe_export_text(social.responsible_rg if social else None, "-")],
            ["Telefone", _safe_export_text(social.phone if social else None, "-")],
            [
                "Estado civil",
                _safe_export_text(social.marital_status if social else None, "-"),
            ],
            [
                "Profissão",
                _safe_export_text(social.profession if social else None, "-"),
            ],
            [
                "Moradores",
                _safe_export_text(social.household_members if social else None, "-"),
            ],
            [
                "Renda familiar",
                _safe_export_text(social.family_income if social else None, "-"),
            ],
            [
                "Forma de ocupação",
                _safe_export_text(social.occupation_type if social else None, "-"),
            ],
            [
                "Documento de posse",
                _safe_export_text(social.possession_document if social else None, "-"),
            ],
            ["Conflito", _yes_no(social.has_conflict) if social else "-"],
        ],
    )

    add_section(
        "4. Cadastro físico",
        [
            ["Campo", "Informação"],
            [
                "Tipo do imóvel",
                _safe_export_text(physical.property_type if physical else None, "-"),
            ],
            [
                "Uso do imóvel",
                _safe_export_text(physical.property_use if physical else None, "-"),
            ],
            [
                "Paredes",
                _safe_export_text(physical.wall_material if physical else None, "-"),
            ],
            [
                "Cobertura",
                _safe_export_text(physical.roof_type if physical else None, "-"),
            ],
            ["Piso", _safe_export_text(physical.floor_type if physical else None, "-")],
            [
                "Pavimentos",
                _safe_export_text(physical.floors if physical else None, "-"),
            ],
            ["Cômodos", _safe_export_text(physical.rooms if physical else None, "-")],
            [
                "Banheiros",
                _safe_export_text(physical.bathrooms if physical else None, "-"),
            ],
            ["Água", _yes_no(physical.has_water) if physical else "-"],
            ["Energia", _yes_no(physical.has_energy) if physical else "-"],
            ["Esgoto", _yes_no(physical.has_sewage) if physical else "-"],
            ["Área de risco", _yes_no(physical.risk_area) if physical else "-"],
            ["Sujeito à inundação", _yes_no(physical.flood_prone) if physical else "-"],
        ],
    )

    add_section(
        "5. Situação documental",
        [
            ["Campo", "Informação"],
            ["Total de documentos", str(documents_count)],
            ["Documentos validados", str(validated_documents_count)],
            [
                "Status documental",
                (
                    "Validado"
                    if documents_count > 0
                    and documents_count == validated_documents_count
                    else "Pendente"
                ),
            ],
        ],
    )

    story.append(Spacer(1, 0.3 * cm))
    story.append(
        Paragraph(
            "Documento gerado automaticamente pelo BIOME REURB para subsidiar "
            "a análise técnico-cadastral e documental do lote. Esta ficha não "
            "substitui planta topográfica, memorial descritivo ou ato administrativo "
            "formal de regularização fundiária.",
            normal,
        )
    )

    doc.build(story)

    buffer.seek(0)
    return buffer.read()


def _load_lot_geojson_geometry(value) -> dict | None:
    """
    Normaliza a geometria armazenada no lote para um objeto GeoJSON Geometry.

    Aceita:
    - string JSON;
    - dict;
    - Feature;
    - FeatureCollection;
    - GeometryCollection;
    - Polygon;
    - MultiPolygon.
    """
    if value is None:
        return None

    data = value

    if isinstance(value, str):
        text = value.strip()

        if not text:
            return None

        try:
            data = json.loads(text)
        except Exception:
            return None

    if not isinstance(data, dict):
        return None

    geo_type = data.get("type")

    if geo_type == "Feature":
        geometry = data.get("geometry")

        if isinstance(geometry, dict):
            return _load_lot_geojson_geometry(geometry)

        return None

    if geo_type == "FeatureCollection":
        features = data.get("features")

        if isinstance(features, list):
            for feature in features:
                geometry = _load_lot_geojson_geometry(feature)

                if geometry is not None:
                    return geometry

        return None

    if geo_type == "GeometryCollection":
        geometries = data.get("geometries")

        if isinstance(geometries, list):
            for geometry in geometries:
                normalized = _load_lot_geojson_geometry(geometry)

                if normalized is not None:
                    return normalized

        return None

    if geo_type in {"Polygon", "MultiPolygon"}:
        coordinates = data.get("coordinates")

        if not coordinates:
            return None

        return data

    return None


def _geometry_outer_rings(geometry: dict) -> list[list[list[float]]]:
    """
    Retorna anéis externos de Polygon/MultiPolygon em coordenadas [lon, lat].
    """
    geom_type = geometry.get("type")
    coordinates = geometry.get("coordinates")

    if not coordinates:
        return []

    if geom_type == "Polygon":
        if coordinates and coordinates[0]:
            return [coordinates[0]]

    if geom_type == "MultiPolygon":
        rings = []

        for polygon in coordinates:
            if polygon and polygon[0]:
                rings.append(polygon[0])

        return rings

    return []


def _point_coordinates(geometry: dict) -> tuple[float, float] | None:
    if geometry.get("type") != "Point":
        return None

    coordinates = geometry.get("coordinates")

    if not isinstance(coordinates, list) or len(coordinates) < 2:
        return None

    try:
        lon = float(coordinates[0])
        lat = float(coordinates[1])
        return lon, lat
    except Exception:
        return None


def _kml_properties_description(properties: dict) -> str:
    rows = []

    for key, value in properties.items():
        if value is None:
            continue

        rows.append(
            f"<tr><td><b>{html.escape(str(key))}</b></td>"
            f"<td>{html.escape(str(value))}</td></tr>"
        )

    return "<![CDATA[<table>" + "".join(rows) + "</table>]]>"


def _feature_collection_to_kml(
    feature_collection: dict,
    *,
    name: str,
) -> str:
    placemarks: list[str] = []

    for feature in feature_collection.get("features", []):
        geometry = feature.get("geometry") or {}
        properties = feature.get("properties") or {}

        placemark_name = (
            properties.get("codigo_lote")
            or properties.get("codigo_selo")
            or properties.get("name")
            or "Feição"
        )

        description = _kml_properties_description(properties)

        if geometry.get("type") in {"Polygon", "MultiPolygon"}:
            rings = _geometry_outer_rings(geometry)

            for ring_index, ring in enumerate(rings, start=1):
                coords = []

                for coord in ring:
                    if not isinstance(coord, (list, tuple)) or len(coord) < 2:
                        continue

                    coords.append(f"{coord[0]},{coord[1]},0")

                if not coords:
                    continue

                placemarks.append(f"""
                    <Placemark>
                      <name>{html.escape(str(placemark_name))}</name>
                      <description>{description}</description>
                      <Polygon>
                        <outerBoundaryIs>
                          <LinearRing>
                            <coordinates>
                              {" ".join(coords)}
                            </coordinates>
                          </LinearRing>
                        </outerBoundaryIs>
                      </Polygon>
                    </Placemark>
                    """)

        elif geometry.get("type") == "Point":
            point = _point_coordinates(geometry)

            if point:
                lon, lat = point

                placemarks.append(f"""
                    <Placemark>
                      <name>{html.escape(str(placemark_name))}</name>
                      <description>{description}</description>
                      <Point>
                        <coordinates>{lon},{lat},0</coordinates>
                      </Point>
                    </Placemark>
                    """)

    return f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document>
    <name>{html.escape(name)}</name>
    {''.join(placemarks)}
  </Document>
</kml>
"""


def _normalize_dbf_field_name(name: str, used: set[str]) -> str:
    base = "".join(ch for ch in name.upper() if ch.isalnum() or ch == "_")[:10]

    if not base:
        base = "CAMPO"

    candidate = base
    index = 1

    while candidate in used:
        suffix = str(index)
        candidate = f"{base[:10 - len(suffix)]}{suffix}"
        index += 1

    used.add(candidate)
    return candidate


def _feature_collection_to_shapefile_zip(
    feature_collection: dict,
    *,
    name: str,
    geometry_type: str,
) -> bytes:
    """
    Gera um ZIP em memória contendo SHP/SHX/DBF/PRJ.
    geometry_type: "polygon" ou "point".
    """
    features = feature_collection.get("features", [])

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        shp_base = temp_path / name

        if geometry_type == "polygon":
            writer = shapefile.Writer(str(shp_base), shapeType=shapefile.POLYGON)
        elif geometry_type == "point":
            writer = shapefile.Writer(str(shp_base), shapeType=shapefile.POINT)
        else:
            raise ValueError("Tipo de geometria não suportado para SHP.")

        property_keys: list[str] = []
        used_fields: set[str] = set()
        field_map: dict[str, str] = {}

        for feature in features:
            properties = feature.get("properties") or {}

            for key in properties.keys():
                if key not in property_keys:
                    property_keys.append(key)

        for key in property_keys:
            field_name = _normalize_dbf_field_name(key, used_fields)
            field_map[key] = field_name
            writer.field(field_name, "C", size=254)

        if not property_keys:
            writer.field("ID", "C", size=80)

        for feature_index, feature in enumerate(features, start=1):
            geometry = feature.get("geometry") or {}
            properties = feature.get("properties") or {}

            if geometry_type == "polygon":
                rings = _geometry_outer_rings(geometry)

                if not rings:
                    continue

                parts = []

                for ring in rings:
                    part = []

                    for coord in ring:
                        if not isinstance(coord, (list, tuple)) or len(coord) < 2:
                            continue

                        part.append([float(coord[0]), float(coord[1])])

                    if len(part) >= 3:
                        if part[0] != part[-1]:
                            part.append(part[0])

                        parts.append(part)

                if not parts:
                    continue

                writer.poly(parts)

            elif geometry_type == "point":
                point = _point_coordinates(geometry)

                if not point:
                    continue

                lon, lat = point
                writer.point(lon, lat)

            if property_keys:
                writer.record(
                    *[
                        _safe_export_text(properties.get(key))[:254]
                        for key in property_keys
                    ]
                )
            else:
                writer.record(str(feature_index))

        writer.close()

        prj_text = (
            'GEOGCS["SIRGAS 2000",'
            'DATUM["Sistema_de_Referencia_Geocentrico_para_las_AmericaS_2000",'
            'SPHEROID["GRS 1980",6378137,298.257222101]],'
            'PRIMEM["Greenwich",0],'
            'UNIT["degree",0.0174532925199433]]'
        )

        prj_path = temp_path / f"{name}.prj"
        prj_path.write_text(prj_text, encoding="utf-8")

        zip_output = temp_path / f"{name}.zip"

        with zipfile.ZipFile(zip_output, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for ext in [".shp", ".shx", ".dbf", ".prj"]:
                file_path = temp_path / f"{name}{ext}"

                if file_path.exists():
                    zip_file.write(file_path, arcname=file_path.name)

        return zip_output.read_bytes()


def _safe_export_text(value, fallback: str = "") -> str:
    if value is None:
        return fallback

    text = str(value).strip()
    return text if text else fallback


def _safe_filename_slug(value: str) -> str:
    text = value.strip().lower()
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE)
    text = re.sub(r"[\s_-]+", "_", text)
    return text.strip("_") or "projeto_reurb"


def _yes_no(value: bool | None) -> str:
    return "Sim" if value else "Não"


def _status_apto(lot: Lot) -> str:
    if getattr(lot, "is_ready_for_technical_documents", False):
        return "Apto para peças técnicas"

    if getattr(lot, "technical_status", None):
        return str(lot.technical_status)

    if getattr(lot, "lot_review_status", None):
        return str(lot.lot_review_status)

    return getattr(lot, "status", None) or "pendente"


def _find_document_file(file_path: str) -> Path | None:
    raw_path = Path(file_path)

    candidates = [
        raw_path,
        STORAGE_DIR / file_path,
        Path("backend") / file_path,
        STORAGE_DIR / "documents" / raw_path.name,
    ]

    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate

    filename = raw_path.name

    if filename:
        imports_dir = STORAGE_DIR / "imports"

        if imports_dir.exists():
            matches = list(imports_dir.rglob(filename))

            for match in matches:
                if match.exists() and match.is_file():
                    return match

    return None


def _ensure_project_access(
    db: Session,
    *,
    project_id: UUID,
    current_user: User,
) -> Project:
    project = db.query(Project).filter(Project.id == project_id).first()

    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projeto não encontrado.",
        )

    if current_user.is_global_admin:
        return project

    link = (
        db.query(ProjectUser)
        .filter(
            ProjectUser.project_id == project_id,
            ProjectUser.user_id == current_user.id,
            ProjectUser.active == True,  # noqa: E712
        )
        .first()
    )

    if link is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Você não possui acesso a este projeto.",
        )

    return project


def _count(db: Session, model, project_id: UUID) -> int:
    return (
        db.query(func.count(model.id)).filter(model.project_id == project_id).scalar()
        or 0
    )


def _lot_to_response(lot: Lot) -> LotResponse:
    return LotResponse(
        id=str(lot.id),
        project_id=str(lot.project_id),
        code=lot.code,
        block=lot.block,
        area_m2=lot.area_m2,
        perimeter_m=lot.perimeter_m,
        status=lot.status,
        needs_review=lot.needs_review,
        source_file=lot.source_file,
        notes=lot.notes,
        lot_review_status=getattr(lot, "lot_review_status", None),
        technical_status=getattr(lot, "technical_status", None),
        is_ready_for_technical_documents=getattr(
            lot,
            "is_ready_for_technical_documents",
            None,
        ),
        geometry_geojson=getattr(lot, "geometry_geojson", None),
        centroid_latitude=getattr(lot, "centroid_latitude", None),
        centroid_longitude=getattr(lot, "centroid_longitude", None),
        geospatial_source=getattr(lot, "geospatial_source", None),
        geospatial_accuracy_m=getattr(lot, "geospatial_accuracy_m", None),
        revision_notes=getattr(lot, "revision_notes", None),
    )


def _seal_to_response(item: Seal, db: Session | None = None) -> SealResponse:
    lot_code = item.lot_code
    responsible_name = None
    responsible_cpf = None
    phone = None
    property_type = None
    property_use = None
    documents_count = 0
    social_count = 0
    physical_count = 0

    if db is not None:
        lot = None

        if item.lot_id:
            lot = (
                db.query(Lot)
                .filter(
                    Lot.id == item.lot_id,
                    Lot.project_id == item.project_id,
                )
                .first()
            )

            if lot:
                lot_code = lot.code

        social = (
            db.query(SocialRegistration)
            .filter(
                SocialRegistration.project_id == item.project_id,
                SocialRegistration.seal_code == item.seal_code,
            )
            .first()
        )

        if social:
            responsible_name = social.responsible_name
            responsible_cpf = social.responsible_cpf
            phone = social.phone

        physical = (
            db.query(PhysicalRegistration)
            .filter(
                PhysicalRegistration.project_id == item.project_id,
                PhysicalRegistration.seal_code == item.seal_code,
            )
            .first()
        )

        if physical:
            property_type = physical.property_type
            property_use = physical.property_use

        social_count = (
            db.query(SocialRegistration)
            .filter(
                SocialRegistration.project_id == item.project_id,
                SocialRegistration.seal_code == item.seal_code,
            )
            .count()
        )

        physical_count = (
            db.query(PhysicalRegistration)
            .filter(
                PhysicalRegistration.project_id == item.project_id,
                PhysicalRegistration.seal_code == item.seal_code,
            )
            .count()
        )

        documents_count = (
            db.query(Document)
            .filter(
                Document.project_id == item.project_id,
                Document.seal_code == item.seal_code,
            )
            .count()
        )

    return SealResponse(
        id=str(item.id),
        project_id=str(item.project_id),
        lot_id=str(item.lot_id) if item.lot_id else None,
        seal_code=item.seal_code,
        lot_code=lot_code,
        situation=item.situation,
        geo_link_status=item.geo_link_status,
        needs_rtk_validation=item.needs_rtk_validation,
        geospatial_note=item.geospatial_note,
        latitude=item.latitude,
        longitude=item.longitude,
        gps_accuracy=item.gps_accuracy,
        responsible_name=responsible_name,
        responsible_cpf=responsible_cpf,
        phone=phone,
        property_type=property_type,
        property_use=property_use,
        social_count=social_count,
        physical_count=physical_count,
        documents_count=documents_count,
    )


def _social_to_response(
    item: SocialRegistration,
    db: Session | None = None,
) -> SocialRegistrationResponse:
    lot_id = None
    lot_code = None
    documents_count = 0

    if db is not None:
        seal = None

        if item.seal_id:
            seal = (
                db.query(Seal)
                .filter(
                    Seal.id == item.seal_id,
                    Seal.project_id == item.project_id,
                )
                .first()
            )

        if seal is None:
            seal = (
                db.query(Seal)
                .filter(
                    Seal.project_id == item.project_id,
                    Seal.seal_code == item.seal_code,
                )
                .first()
            )

        if seal:
            lot_id = str(seal.lot_id) if seal.lot_id else None
            lot_code = seal.lot_code

            if seal.lot_id:
                lot = (
                    db.query(Lot)
                    .filter(
                        Lot.id == seal.lot_id,
                        Lot.project_id == item.project_id,
                    )
                    .first()
                )

                if lot:
                    lot_code = lot.code

        documents_count = (
            db.query(Document)
            .filter(
                Document.project_id == item.project_id,
                Document.social_registration_id == item.id,
            )
            .count()
        )

    return SocialRegistrationResponse(
        id=str(item.id),
        project_id=str(item.project_id),
        seal_id=str(item.seal_id) if item.seal_id else None,
        seal_code=item.seal_code,
        lot_id=lot_id,
        lot_code=lot_code,
        responsible_name=item.responsible_name,
        responsible_cpf=item.responsible_cpf,
        responsible_rg=item.responsible_rg,
        issuing_agency=item.issuing_agency,
        phone=item.phone,
        marital_status=item.marital_status,
        profession=item.profession,
        household_members=item.household_members,
        family_income=item.family_income,
        receives_social_program=item.receives_social_program,
        social_program=item.social_program,
        occupation_years=item.occupation_years,
        occupation_type=item.occupation_type,
        possession_document=item.possession_document,
        owns_other_property=item.owns_other_property,
        has_conflict=item.has_conflict,
        notes=item.notes,
        documents_count=documents_count,
    )


def _physical_to_response(
    item: PhysicalRegistration,
    db: Session | None = None,
) -> PhysicalRegistrationResponse:
    lot_id = None
    lot_code = None

    if db is not None:
        seal = None

        if item.seal_id:
            seal = (
                db.query(Seal)
                .filter(
                    Seal.id == item.seal_id,
                    Seal.project_id == item.project_id,
                )
                .first()
            )

        if seal is None:
            seal = (
                db.query(Seal)
                .filter(
                    Seal.project_id == item.project_id,
                    Seal.seal_code == item.seal_code,
                )
                .first()
            )

        if seal:
            lot_id = str(seal.lot_id) if seal.lot_id else None
            lot_code = seal.lot_code

            if seal.lot_id:
                lot = (
                    db.query(Lot)
                    .filter(
                        Lot.id == seal.lot_id,
                        Lot.project_id == item.project_id,
                    )
                    .first()
                )

                if lot:
                    lot_code = lot.code

    return PhysicalRegistrationResponse(
        id=str(item.id),
        project_id=str(item.project_id),
        seal_id=str(item.seal_id) if item.seal_id else None,
        seal_code=item.seal_code,
        lot_id=lot_id,
        lot_code=lot_code,
        property_type=item.property_type,
        property_use=item.property_use,
        wall_material=item.wall_material,
        roof_type=item.roof_type,
        floor_type=item.floor_type,
        floors=item.floors,
        rooms=item.rooms,
        bathrooms=item.bathrooms,
        has_energy=item.has_energy,
        has_water=item.has_water,
        has_sewage=item.has_sewage,
        has_bathroom=item.has_bathroom,
        habitability_condition=item.habitability_condition,
        risk_area=item.risk_area,
        flood_prone=item.flood_prone,
        notes=item.notes,
    )


DOCUMENT_STORAGE_DIR = Path("storage/documents")
IMPORT_STORAGE_DIR = Path("storage/imports")


def _safe_filename(filename: str) -> str:
    filename = filename.replace("\\", "/").split("/")[-1].strip()

    if not filename:
        return "documento"

    blocked = {"..", ".", "/", "\\"}

    for item in blocked:
        filename = filename.replace(item, "")

    return filename or "documento"


def _extract_file_extension(filename: str, content_type: str | None = None) -> str:
    suffix = Path(filename).suffix.lower()

    if suffix:
        return suffix

    if content_type:
        guessed = mimetypes.guess_extension(content_type.split(";")[0].strip())

        if guessed:
            return guessed

    return ""


def _resolve_document_path(document: Document) -> Path | None:
    """
    Resolve documentos novos salvos no storage/documents e documentos antigos/importados
    vindos do mobile em storage/imports/**/extracted.
    """
    raw_path = (document.file_path or "").strip()

    if not raw_path:
        return None

    candidates: list[Path] = []

    direct = Path(raw_path)

    if direct.is_absolute():
        candidates.append(direct)
    else:
        candidates.append(Path(raw_path))
        candidates.append(Path("storage") / raw_path)

    filename = Path(raw_path.replace("\\", "/")).name

    if filename:
        candidates.append(DOCUMENT_STORAGE_DIR / str(document.project_id) / filename)

        if IMPORT_STORAGE_DIR.exists():
            candidates.extend(IMPORT_STORAGE_DIR.rglob(filename))

    for candidate in candidates:
        try:
            if candidate.exists() and candidate.is_file():
                return candidate
        except Exception:
            continue

    return None


def _copy_mobile_document_to_project_storage(
    *,
    project_id: UUID,
    source_path: Path,
    fallback_filename: str,
) -> str:
    """
    Quando o documento veio do mobile/imports/extracted, copia para storage/documents
    para padronizar abertura futura.
    """
    DOCUMENT_STORAGE_DIR.mkdir(parents=True, exist_ok=True)

    project_dir = DOCUMENT_STORAGE_DIR / str(project_id)
    project_dir.mkdir(parents=True, exist_ok=True)

    original_name = _safe_filename(fallback_filename or source_path.name)
    suffix = Path(original_name).suffix or source_path.suffix

    stored_name = f"{uuid.uuid4()}{suffix}"
    target_path = project_dir / stored_name

    shutil.copyfile(source_path, target_path)

    return str(target_path)


def _document_to_response(item: Document) -> DocumentResponse:
    return DocumentResponse(
        id=str(item.id),
        project_id=str(item.project_id),
        lot_id=str(item.lot_id) if getattr(item, "lot_id", None) else None,
        seal_id=str(item.seal_id) if item.seal_id else None,
        social_registration_id=(
            str(item.social_registration_id) if item.social_registration_id else None
        ),
        seal_code=item.seal_code,
        document_type=item.document_type,
        file_path=item.file_path,
        original_filename=getattr(item, "original_filename", None),
        stored_filename=getattr(item, "stored_filename", None),
        mime_type=getattr(item, "mime_type", None),
        file_size_bytes=getattr(item, "file_size_bytes", None),
        notes=item.notes,
        validated=item.validated,
        document_status=getattr(item, "document_status", None),
        validation_notes=getattr(item, "validation_notes", None),
        validated_at=(
            item.validated_at.isoformat()
            if getattr(item, "validated_at", None)
            else None
        ),
        validated_by_user_id=(
            str(item.validated_by_user_id)
            if getattr(item, "validated_by_user_id", None)
            else None
        ),
    )


def _ensure_administrative_access(
    current_user: User,
) -> None:
    if not current_user.is_global_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=(
                "Somente um administrador global pode excluir "
                "selagens pelo painel web."
            ),
        )


@router.get("/dashboard", response_model=ProjectDashboardResponse)
def get_project_dashboard(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ProjectDashboardResponse:
    project = _ensure_project_access(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    total_lots = _count(db, Lot, project_id)
    total_seals = _count(db, Seal, project_id)
    total_social = _count(db, SocialRegistration, project_id)
    total_physical = _count(db, PhysicalRegistration, project_id)
    total_documents = _count(db, Document, project_id)

    lots_without_seal = (
        db.query(func.count(Lot.id))
        .outerjoin(Seal, Seal.lot_id == Lot.id)
        .filter(
            Lot.project_id == project_id,
            Seal.id.is_(None),
        )
        .scalar()
        or 0
    )

    seals_without_social = (
        db.query(func.count(Seal.id))
        .outerjoin(SocialRegistration, SocialRegistration.seal_id == Seal.id)
        .filter(
            Seal.project_id == project_id,
            SocialRegistration.id.is_(None),
        )
        .scalar()
        or 0
    )

    seals_without_physical = (
        db.query(func.count(Seal.id))
        .outerjoin(PhysicalRegistration, PhysicalRegistration.seal_id == Seal.id)
        .filter(
            Seal.project_id == project_id,
            PhysicalRegistration.id.is_(None),
        )
        .scalar()
        or 0
    )

    social_without_documents = (
        db.query(func.count(SocialRegistration.id))
        .outerjoin(Document, Document.social_registration_id == SocialRegistration.id)
        .filter(
            SocialRegistration.project_id == project_id,
            Document.id.is_(None),
        )
        .scalar()
        or 0
    )

    seals_needing_rtk = (
        db.query(func.count(Seal.id))
        .filter(
            Seal.project_id == project_id,
            Seal.needs_rtk_validation == True,  # noqa: E712
        )
        .scalar()
        or 0
    )

    return ProjectDashboardResponse(
        project_id=str(project.id),
        project_name=project.name,
        total_lots=total_lots,
        total_seals=total_seals,
        total_social_registrations=total_social,
        total_physical_registrations=total_physical,
        total_documents=total_documents,
        lots_without_seal=lots_without_seal,
        seals_without_social=seals_without_social,
        seals_without_physical=seals_without_physical,
        social_without_documents=social_without_documents,
        seals_needing_rtk=seals_needing_rtk,
    )


@router.get("/lots", response_model=list[LotResponse])
def list_project_lots(
    project_id: UUID,
    search: str | None = Query(default=None),
    needs_review: bool | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[LotResponse]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    query = db.query(Lot).filter(Lot.project_id == project_id)

    if search:
        like = f"%{search}%"
        query = query.filter(
            Lot.code.ilike(like) | Lot.block.ilike(like) | Lot.status.ilike(like)
        )

    if needs_review is not None:
        query = query.filter(Lot.needs_review == needs_review)

    lots = query.order_by(Lot.code.asc()).offset(offset).limit(limit).all()

    return [_lot_to_response(lot) for lot in lots]


@router.get("/seals", response_model=list[SealResponse])
def list_project_seals(
    project_id: UUID,
    search: str | None = Query(default=None),
    needs_rtk_validation: bool | None = Query(default=None),
    geo_link_status: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[SealResponse]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    query = db.query(Seal).filter(Seal.project_id == project_id)

    if search:
        like = f"%{search}%"
        query = query.filter(
            Seal.seal_code.ilike(like)
            | Seal.lot_code.ilike(like)
            | Seal.situation.ilike(like)
        )

    if needs_rtk_validation is not None:
        query = query.filter(Seal.needs_rtk_validation == needs_rtk_validation)

    if geo_link_status:
        query = query.filter(Seal.geo_link_status == geo_link_status)

    seals = query.order_by(Seal.seal_code.asc()).offset(offset).limit(limit).all()

    return [_seal_to_response(item, db=db) for item in seals]


@router.get("/social-registrations", response_model=list[SocialRegistrationResponse])
def list_project_social_registrations(
    project_id: UUID,
    search: str | None = Query(default=None),
    has_conflict: bool | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[SocialRegistrationResponse]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    query = db.query(SocialRegistration).filter(
        SocialRegistration.project_id == project_id
    )

    if search:
        like = f"%{search}%"
        query = query.filter(
            SocialRegistration.responsible_name.ilike(like)
            | SocialRegistration.responsible_cpf.ilike(like)
            | SocialRegistration.seal_code.ilike(like)
        )

    if has_conflict is not None:
        query = query.filter(SocialRegistration.has_conflict == has_conflict)

    items = (
        query.order_by(SocialRegistration.responsible_name.asc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    return [_social_to_response(item, db=db) for item in items]


@router.get(
    "/physical-registrations", response_model=list[PhysicalRegistrationResponse]
)
def list_project_physical_registrations(
    project_id: UUID,
    search: str | None = Query(default=None),
    risk_area: bool | None = Query(default=None),
    flood_prone: bool | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[PhysicalRegistrationResponse]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    query = db.query(PhysicalRegistration).filter(
        PhysicalRegistration.project_id == project_id
    )

    if search:
        like = f"%{search}%"
        query = query.filter(
            PhysicalRegistration.seal_code.ilike(like)
            | PhysicalRegistration.property_type.ilike(like)
            | PhysicalRegistration.property_use.ilike(like)
        )

    if risk_area is not None:
        query = query.filter(PhysicalRegistration.risk_area == risk_area)

    if flood_prone is not None:
        query = query.filter(PhysicalRegistration.flood_prone == flood_prone)

    items = (
        query.order_by(PhysicalRegistration.seal_code.asc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    return [_physical_to_response(item, db=db) for item in items]


@router.get("/documents", response_model=list[DocumentResponse])
def list_project_documents(
    project_id: UUID,
    search: str | None = Query(default=None),
    document_type: str | None = Query(default=None),
    validated: bool | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[DocumentResponse]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    query = db.query(Document).filter(Document.project_id == project_id)

    if search:
        like = f"%{search}%"
        query = query.filter(
            Document.seal_code.ilike(like)
            | Document.document_type.ilike(like)
            | Document.file_path.ilike(like)
        )

    if document_type:
        query = query.filter(Document.document_type == document_type)

    if validated is not None:
        query = query.filter(Document.validated == validated)

    items = query.order_by(Document.created_at.desc()).offset(offset).limit(limit).all()

    return [_document_to_response(item) for item in items]


def _build_lot_pending_flags(
    *,
    lot: Lot,
    seal: Seal | None,
    social: SocialRegistration | None,
    physical: PhysicalRegistration | None,
    documents_count: int,
) -> list[str]:
    flags: list[str] = []

    if not getattr(lot, "geometry_geojson", None) and lot.geom is None:
        flags.append("sem_geometria")

    if seal is None:
        flags.append("sem_selagem")

    if social is None:
        flags.append("sem_cadastro_social")

    if physical is None:
        flags.append("sem_cadastro_fisico")

    if documents_count <= 0:
        flags.append("sem_documentos")

    if seal and seal.needs_rtk_validation:
        flags.append("necessita_rtk")

    if social and social.has_conflict:
        flags.append("conflito_cadastral")

    if physical and physical.risk_area:
        flags.append("area_de_risco")

    if physical and physical.flood_prone:
        flags.append("sujeito_inundacao")

    return flags


DOCUMENT_STORAGE_DIR = Path("storage/documents")


def _get_lot_or_404(
    db: Session,
    *,
    project_id: UUID,
    lot_id: UUID,
) -> Lot:
    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    return lot


def _get_lot_seal(
    db: Session,
    *,
    project_id: UUID,
    lot_id: UUID,
) -> Seal | None:
    return (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot_id,
        )
        .order_by(Seal.created_at.asc())
        .first()
    )


def _get_lot_social_registration(
    db: Session,
    *,
    project_id: UUID,
    seal_code: str | None,
) -> SocialRegistration | None:
    if not seal_code:
        return None

    return (
        db.query(SocialRegistration)
        .filter(
            SocialRegistration.project_id == project_id,
            SocialRegistration.seal_code == seal_code,
        )
        .first()
    )


def _lot_document_to_response(document: Document) -> LotDocumentResponse:
    return LotDocumentResponse(
        id=str(document.id),
        project_id=str(document.project_id),
        lot_id=str(document.lot_id) if getattr(document, "lot_id", None) else None,
        seal_id=str(document.seal_id) if document.seal_id else None,
        social_registration_id=(
            str(document.social_registration_id)
            if document.social_registration_id
            else None
        ),
        seal_code=document.seal_code,
        document_type=document.document_type,
        file_path=document.file_path,
        original_filename=getattr(document, "original_filename", None),
        stored_filename=getattr(document, "stored_filename", None),
        mime_type=getattr(document, "mime_type", None),
        file_size_bytes=getattr(document, "file_size_bytes", None),
        notes=document.notes,
        validated=document.validated,
        document_status=getattr(document, "document_status", "pendente"),
        validation_notes=getattr(document, "validation_notes", None),
        validated_at=(
            document.validated_at.isoformat()
            if getattr(document, "validated_at", None)
            else None
        ),
        validated_by_user_id=(
            str(document.validated_by_user_id)
            if getattr(document, "validated_by_user_id", None)
            else None
        ),
    )


def _query_lot_documents(
    db: Session,
    *,
    project_id: UUID,
    lot: Lot,
) -> list[Document]:
    seal = _get_lot_seal(db, project_id=project_id, lot_id=lot.id)

    query = db.query(Document).filter(Document.project_id == project_id)

    if seal is not None:
        query = query.filter(
            (Document.lot_id == lot.id) | (Document.seal_code == seal.seal_code)
        )
    else:
        query = query.filter(Document.lot_id == lot.id)

    return query.order_by(Document.created_at.desc()).all()


def _save_lot_document_upload(
    *,
    project_id: UUID,
    lot_id: UUID,
    file: UploadFile,
) -> tuple[Path, str, str, int]:
    original_filename = file.filename or "documento"
    suffix = Path(original_filename).suffix.lower()

    stored_filename = f"{uuid.uuid4().hex}{suffix}"

    target_dir = DOCUMENT_STORAGE_DIR / str(project_id) / str(lot_id)
    target_dir.mkdir(parents=True, exist_ok=True)

    target_path = target_dir / stored_filename

    with target_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    file_size = target_path.stat().st_size

    return target_path, original_filename, stored_filename, file_size


@router.get("/map", response_model=ProjectMapResponse)
def get_project_map(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ProjectMapResponse:
    project = _ensure_project_access(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    lots = (
        db.query(Lot)
        .filter(Lot.project_id == project_id)
        .order_by(Lot.code.asc())
        .all()
    )

    result_lots: list[ProjectMapLotResponse] = []

    ready_lots = 0
    inconsistent_lots = 0
    lots_without_geometry = 0
    lots_without_seal = 0

    for lot in lots:
        seal = (
            db.query(Seal)
            .filter(
                Seal.project_id == project_id,
                Seal.lot_id == lot.id,
            )
            .order_by(Seal.created_at.asc())
            .first()
        )

        seal = (
            db.query(Seal)
            .filter(
                Seal.project_id == project_id,
                Seal.lot_id == lot.id,
            )
            .order_by(Seal.created_at.asc())
            .first()
        )

        social = None
        physical = None
        documents_count = 0

        if seal is not None:
            social = (
                db.query(SocialRegistration)
                .filter(
                    SocialRegistration.project_id == project_id,
                    SocialRegistration.seal_code == seal.seal_code,
                )
                .first()
            )

            physical = (
                db.query(PhysicalRegistration)
                .filter(
                    PhysicalRegistration.project_id == project_id,
                    PhysicalRegistration.seal_code == seal.seal_code,
                )
                .first()
            )

            documents_count = (
                db.query(func.count(Document.id))
                .filter(
                    Document.project_id == project_id,
                    Document.seal_code == seal.seal_code,
                )
                .scalar()
                or 0
            )

        pending_flags = _build_lot_pending_flags(
            lot=lot,
            seal=seal,
            social=social,
            physical=physical,
            documents_count=documents_count,
        )

        if getattr(lot, "is_ready_for_technical_documents", False):
            ready_lots += 1

        if getattr(lot, "lot_review_status", "") == "inconsistente":
            inconsistent_lots += 1

        if "sem_geometria" in pending_flags:
            lots_without_geometry += 1

        if seal is None:
            lots_without_seal += 1

        result_lots.append(
            ProjectMapLotResponse(
                id=str(lot.id),
                code=lot.code,
                block=lot.block,
                area_m2=lot.area_m2,
                perimeter_m=lot.perimeter_m,
                status=lot.status,
                needs_review=lot.needs_review,
                lot_review_status=getattr(
                    lot,
                    "lot_review_status",
                    "preliminar",
                ),
                technical_status=getattr(
                    lot,
                    "technical_status",
                    "sem_geometria",
                ),
                is_ready_for_technical_documents=getattr(
                    lot,
                    "is_ready_for_technical_documents",
                    False,
                ),
                geometry_geojson=getattr(lot, "geometry_geojson", None),
                centroid_latitude=getattr(lot, "centroid_latitude", None),
                centroid_longitude=getattr(lot, "centroid_longitude", None),
                geospatial_source=getattr(lot, "geospatial_source", None),
                geospatial_accuracy_m=getattr(lot, "geospatial_accuracy_m", None),
                revision_notes=getattr(lot, "revision_notes", None),
                seal=(
                    ProjectMapSealResponse(
                        id=str(seal.id),
                        seal_code=seal.seal_code,
                        lot_code=seal.lot_code,
                        situation=seal.situation,
                        geo_link_status=seal.geo_link_status,
                        needs_rtk_validation=seal.needs_rtk_validation,
                        geospatial_note=seal.geospatial_note,
                        latitude=seal.latitude,
                        longitude=seal.longitude,
                        gps_accuracy=seal.gps_accuracy,
                    )
                    if seal
                    else None
                ),
                social=(
                    ProjectMapSocialResponse(
                        id=str(social.id),
                        responsible_name=social.responsible_name,
                        responsible_cpf=social.responsible_cpf,
                        phone=social.phone,
                        household_members=social.household_members,
                        family_income=social.family_income,
                        has_conflict=social.has_conflict,
                    )
                    if social
                    else None
                ),
                physical=(
                    ProjectMapPhysicalResponse(
                        id=str(physical.id),
                        property_type=physical.property_type,
                        property_use=physical.property_use,
                        wall_material=physical.wall_material,
                        roof_type=physical.roof_type,
                        floor_type=physical.floor_type,
                        rooms=physical.rooms,
                        bathrooms=physical.bathrooms,
                        has_energy=physical.has_energy,
                        has_water=physical.has_water,
                        has_sewage=physical.has_sewage,
                        has_bathroom=physical.has_bathroom,
                        habitability_condition=physical.habitability_condition,
                        risk_area=physical.risk_area,
                        flood_prone=physical.flood_prone,
                    )
                    if physical
                    else None
                ),
                documents_count=documents_count,
                pending_flags=pending_flags,
            )
        )

    seals_without_lot_items = (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id.is_(None),
        )
        .order_by(Seal.seal_code.asc())
        .all()
    )

    seals_without_lot = [
        ProjectMapSealWithoutLotResponse(
            id=str(seal.id),
            seal_code=seal.seal_code,
            lot_code=seal.lot_code,
            latitude=seal.latitude,
            longitude=seal.longitude,
            geo_link_status=seal.geo_link_status,
            needs_rtk_validation=seal.needs_rtk_validation,
        )
        for seal in seals_without_lot_items
    ]

    return ProjectMapResponse(
        project=ProjectMapProjectResponse(
            id=str(project.id),
            name=project.name,
            municipality=project.municipality,
            state=project.state,
            neighborhood=project.neighborhood,
            reurb_type=project.reurb_type,
            status=project.status,
        ),
        summary=ProjectMapSummaryResponse(
            total_lots=len(result_lots),
            ready_lots=ready_lots,
            pending_lots=len(result_lots) - ready_lots,
            inconsistent_lots=inconsistent_lots,
            lots_without_geometry=lots_without_geometry,
            lots_without_seal=lots_without_seal,
            seals_without_lot=len(seals_without_lot),
        ),
        lots=result_lots,
        seals_without_lot=seals_without_lot,
    )


@router.patch("/lots/{lot_id}/review", response_model=LotReviewResponse)
def update_lot_review(
    project_id: UUID,
    lot_id: UUID,
    payload: LotReviewUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> LotReviewResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    allowed_review_status = {
        "preliminar",
        "em_revisao",
        "inconsistente",
        "apto",
    }

    if payload.lot_review_status not in allowed_review_status:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Status de revisão inválido.",
        )

    old_data = {
        "lot_review_status": getattr(lot, "lot_review_status", None),
        "technical_status": getattr(lot, "technical_status", None),
        "is_ready_for_technical_documents": getattr(
            lot,
            "is_ready_for_technical_documents",
            None,
        ),
        "revision_notes": getattr(lot, "revision_notes", None),
    }

    lot.lot_review_status = payload.lot_review_status

    if payload.technical_status is not None:
        lot.technical_status = payload.technical_status
    else:
        if payload.lot_review_status == "apto":
            lot.technical_status = "apto_para_pecas"
        elif payload.lot_review_status == "inconsistente":
            lot.technical_status = "inconsistente"
        elif payload.lot_review_status == "em_revisao":
            lot.technical_status = "em_revisao"
        else:
            lot.technical_status = "sem_geometria"

    if payload.is_ready_for_technical_documents is not None:
        lot.is_ready_for_technical_documents = payload.is_ready_for_technical_documents
    else:
        lot.is_ready_for_technical_documents = payload.lot_review_status == "apto"

    lot.revision_notes = payload.revision_notes

    if lot.is_ready_for_technical_documents:
        lot.approved_at = datetime.now(timezone.utc)
        lot.approved_by_user_id = current_user.id
    else:
        lot.approved_at = None
        lot.approved_by_user_id = None

    db.commit()
    db.refresh(lot)

    register_audit_log(
        db,
        user=current_user,
        action="UPDATE",
        entity_type="lot_review",
        entity_id=lot.id,
        project_id=project_id,
        description=f"Atualizou revisão técnica do lote {lot.code}.",
        old_data=old_data,
        new_data={
            "lot_review_status": lot.lot_review_status,
            "technical_status": lot.technical_status,
            "is_ready_for_technical_documents": lot.is_ready_for_technical_documents,
            "revision_notes": lot.revision_notes,
        },
        request=request,
        severity="WARNING",
    )

    return LotReviewResponse(
        id=str(lot.id),
        project_id=str(lot.project_id),
        code=lot.code,
        lot_review_status=lot.lot_review_status,
        technical_status=lot.technical_status,
        is_ready_for_technical_documents=lot.is_ready_for_technical_documents,
        revision_notes=lot.revision_notes,
    )


@router.post("/geospatial/lots/import")
async def import_project_lot_geometries(
    project_id: UUID,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = _ensure_project_access(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    filename = file.filename or ""

    allowed = filename.lower().endswith(
        (
            ".geojson",
            ".json",
            ".kml",
            ".zip",
            ".shp",
        )
    )

    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato inválido. Envie .geojson, .json, .kml ou .zip contendo Shapefile.",
        )

    try:
        file_path, original_filename = await save_geospatial_upload(file)

        result = import_lot_geometries_to_project(
            db,
            project_id=project_id,
            file_path=file_path,
            original_filename=original_filename,
        )

        register_audit_log(
            db,
            user=current_user,
            action="CREATE",
            entity_type="geospatial_lot_import",
            entity_id=project_id,
            project_id=project_id,
            description=f"Importou arquivo geoespacial de lotes no projeto {project.name}: {original_filename}.",
            new_data={
                "filename": original_filename,
                "stored_path": str(file_path),
                **result,
            },
            request=request,
            severity="INFO",
        )

        return {
            "status": "imported",
            "filename": original_filename,
            **result,
        }

    except Exception as exc:
        db.rollback()

        error_message = str(exc)
        project_name = getattr(project, "name", str(project_id))

        try:
            register_audit_log(
                db,
                user=current_user,
                action="ERROR",
                entity_type="geospatial_lot_import",
                entity_id=project_id,
                project_id=project_id,
                description=(
                    f"Erro ao importar arquivo geoespacial de lotes no projeto "
                    f"{project_name}: {filename}."
                ),
                new_data={
                    "filename": filename,
                    "error": error_message,
                },
                request=request,
                severity="ERROR",
            )
        except Exception:
            db.rollback()

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao importar arquivo geoespacial: {error_message}",
        )


@router.get("/lots/{lot_id}/delete-check")
def check_lot_delete(
    project_id: UUID,
    lot_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seals_count = (
        db.query(func.count(Seal.id))
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
        )
        .scalar()
        or 0
    )

    social_count = (
        db.query(func.count(SocialRegistration.id))
        .join(Seal, Seal.seal_code == SocialRegistration.seal_code)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
            SocialRegistration.project_id == project_id,
        )
        .scalar()
        or 0
    )

    physical_count = (
        db.query(func.count(PhysicalRegistration.id))
        .join(Seal, Seal.seal_code == PhysicalRegistration.seal_code)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
            PhysicalRegistration.project_id == project_id,
        )
        .scalar()
        or 0
    )

    documents_count = (
        db.query(func.count(Document.id))
        .join(Seal, Seal.seal_code == Document.seal_code)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
            Document.project_id == project_id,
        )
        .scalar()
        or 0
    )

    total_links = seals_count + social_count + physical_count + documents_count

    return {
        "can_delete": total_links == 0,
        "lot_id": str(lot.id),
        "lot_code": lot.code,
        "links": {
            "seals": seals_count,
            "social_registrations": social_count,
            "physical_registrations": physical_count,
            "documents": documents_count,
        },
        "message": (
            "Lote pode ser excluído."
            if total_links == 0
            else "Lote possui vínculos ativos. Desvincule a selagem antes de excluir."
        ),
    }


@router.get("/lots/{lot_id}/link-candidates")
def list_lot_link_candidates(
    project_id: UUID,
    lot_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seals = (
        db.query(Seal)
        .filter(Seal.project_id == project_id)
        .order_by(Seal.seal_code.asc())
        .all()
    )

    candidates = []

    for seal in seals:
        social = (
            db.query(SocialRegistration)
            .filter(
                SocialRegistration.project_id == project_id,
                SocialRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

        physical = (
            db.query(PhysicalRegistration)
            .filter(
                PhysicalRegistration.project_id == project_id,
                PhysicalRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

        documents_count = (
            db.query(func.count(Document.id))
            .filter(
                Document.project_id == project_id,
                Document.seal_code == seal.seal_code,
            )
            .scalar()
            or 0
        )

        if seal.lot_id is None:
            link_status = "disponivel"
            can_link = True
            warning = None
        elif seal.lot_id == lot.id:
            link_status = "ja_vinculada_este_lote"
            can_link = False
            warning = "Esta selagem já está vinculada a este lote."
        else:
            link_status = "vinculada_outro_lote"
            can_link = True
            warning = (
                "Esta selagem já está vinculada a outro lote. "
                "Ao confirmar, o vínculo será transferido para este lote."
            )

        code_match = bool(seal.lot_code and lot.code and seal.lot_code == lot.code)

        candidates.append(
            {
                "seal_id": str(seal.id),
                "seal_code": seal.seal_code,
                "lot_code": seal.lot_code,
                "responsible_name": social.responsible_name if social else None,
                "latitude": seal.latitude,
                "longitude": seal.longitude,
                "geo_link_status": seal.geo_link_status,
                "distance_m": None,
                "has_social": social is not None,
                "has_physical": physical is not None,
                "documents_count": documents_count,
                "link_status": link_status,
                "can_link": can_link,
                "code_match": code_match,
                "warning": warning,
            }
        )

    candidates.sort(
        key=lambda item: (
            0 if item["code_match"] else 1,
            0 if item["link_status"] == "disponivel" else 1,
            item["seal_code"],
        )
    )

    return candidates


@router.patch("/lots/{lot_id}/link-seal")
def link_seal_to_lot(
    project_id: UUID,
    lot_id: UUID,
    payload: dict,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    seal_id = payload.get("seal_id")

    if not seal_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Informe seal_id.",
        )

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seal = (
        db.query(Seal)
        .filter(
            Seal.id == seal_id,
            Seal.project_id == project_id,
        )
        .first()
    )

    if seal is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Selagem não encontrada.",
        )

    old_lot_id = seal.lot_id

    seal.lot_id = lot.id
    seal.lot_code = lot.code
    seal.geo_link_status = "confirmado"

    lot.needs_review = True
    lot.technical_status = "pendente"
    lot.lot_review_status = "preliminar"
    lot.is_ready_for_technical_documents = False
    lot.revision_notes = (
        f"Selagem {seal.seal_code} vinculada manualmente ao lote {lot.code}."
    )

    db.commit()
    db.refresh(lot)
    db.refresh(seal)

    register_audit_log(
        db,
        user=current_user,
        action="UPDATE",
        entity_type="lot_seal_link",
        entity_id=lot.id,
        project_id=project_id,
        description=f"Vinculou a selagem {seal.seal_code} ao lote {lot.code}.",
        old_data={
            "seal_id": str(seal.id),
            "old_lot_id": str(old_lot_id) if old_lot_id else None,
        },
        new_data={
            "seal_id": str(seal.id),
            "new_lot_id": str(lot.id),
            "lot_code": lot.code,
        },
        request=request,
        severity="WARNING",
    )

    return {
        "status": "linked",
        "lot_id": str(lot.id),
        "lot_code": lot.code,
        "seal_id": str(seal.id),
        "seal_code": seal.seal_code,
        "old_lot_id": str(old_lot_id) if old_lot_id else None,
    }


@router.patch("/lots/{lot_id}/unlink-seal")
def unlink_seal_from_lot(
    project_id: UUID,
    lot_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seal = (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
        )
        .first()
    )

    if seal is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Este lote não possui selagem vinculada.",
        )

    old_data = {
        "seal_id": str(seal.id),
        "seal_code": seal.seal_code,
        "lot_id": str(lot.id),
        "lot_code": lot.code,
    }

    seal.lot_id = None
    seal.geo_link_status = "nao_validado"

    lot.needs_review = True
    lot.lot_review_status = "preliminar"
    lot.technical_status = "pendente"
    lot.is_ready_for_technical_documents = False
    lot.revision_notes = f"Selagem {seal.seal_code} desvinculada manualmente."

    db.commit()
    db.refresh(lot)
    db.refresh(seal)

    register_audit_log(
        db,
        user=current_user,
        action="UPDATE",
        entity_type="lot_seal_unlink",
        entity_id=lot.id,
        project_id=project_id,
        description=f"Desvinculou a selagem {seal.seal_code} do lote {lot.code}.",
        old_data=old_data,
        new_data={
            "seal_id": str(seal.id),
            "seal_code": seal.seal_code,
            "lot_id": None,
            "geo_link_status": seal.geo_link_status,
        },
        request=request,
        severity="WARNING",
    )

    return {
        "status": "unlinked",
        "lot_id": str(lot.id),
        "lot_code": lot.code,
        "seal_id": str(seal.id),
        "seal_code": seal.seal_code,
    }


@router.delete("/lots/{lot_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_lot(
    project_id: UUID,
    lot_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> None:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    linked_seal = (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
        )
        .first()
    )

    if linked_seal is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Este lote possui selagem/cadastro vinculado. "
                "Desvincule a selagem antes de excluir o lote."
            ),
        )

    old_data = {
        "id": str(lot.id),
        "code": lot.code,
        "area_m2": lot.area_m2,
        "perimeter_m": lot.perimeter_m,
        "source_file": lot.source_file,
        "lot_review_status": lot.lot_review_status,
        "technical_status": lot.technical_status,
    }

    db.delete(lot)
    db.commit()

    register_audit_log(
        db,
        user=current_user,
        action="DELETE",
        entity_type="lot",
        entity_id=lot_id,
        project_id=project_id,
        description=f"Excluiu o lote {old_data['code']}.",
        old_data=old_data,
        request=request,
        severity="CRITICAL",
    )


@router.get(
    "/lots/{lot_id}/delete-check",
    response_model=LotDeleteCheckResponse,
)
def check_lot_before_delete(
    project_id: UUID,
    lot_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> LotDeleteCheckResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seals = (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
        )
        .all()
    )

    seal_codes = [seal.seal_code for seal in seals]

    social_count = 0
    physical_count = 0
    documents_count = 0

    if seal_codes:
        social_count = (
            db.query(func.count(SocialRegistration.id))
            .filter(
                SocialRegistration.project_id == project_id,
                SocialRegistration.seal_code.in_(seal_codes),
            )
            .scalar()
            or 0
        )

        physical_count = (
            db.query(func.count(PhysicalRegistration.id))
            .filter(
                PhysicalRegistration.project_id == project_id,
                PhysicalRegistration.seal_code.in_(seal_codes),
            )
            .scalar()
            or 0
        )

        documents_count = (
            db.query(func.count(Document.id))
            .filter(
                Document.project_id == project_id,
                Document.seal_code.in_(seal_codes),
            )
            .scalar()
            or 0
        )

    links = {
        "seals": len(seals),
        "social_registrations": social_count,
        "physical_registrations": physical_count,
        "documents": documents_count,
    }

    can_delete = all(value == 0 for value in links.values())

    return LotDeleteCheckResponse(
        can_delete=can_delete,
        lot_id=str(lot.id),
        lot_code=lot.code,
        links=links,
        message=(
            "Lote sem vínculos. Pode ser excluído."
            if can_delete
            else "Este lote possui vínculos. Desvincule as selagens/cadastros antes de excluir."
        ),
    )


@router.delete("/lots/{lot_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_lot(
    project_id: UUID,
    lot_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> None:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    linked_seals_count = (
        db.query(func.count(Seal.id))
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
        )
        .scalar()
        or 0
    )

    if linked_seals_count > 0:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Este lote possui selagens vinculadas. "
                "Desvincule os cadastros antes de excluir o lote."
            ),
        )

    old_data = {
        "id": str(lot.id),
        "code": lot.code,
        "block": lot.block,
        "area_m2": lot.area_m2,
        "perimeter_m": lot.perimeter_m,
        "source_file": lot.source_file,
    }

    db.delete(lot)
    db.commit()

    register_audit_log(
        db,
        user=current_user,
        action="DELETE",
        entity_type="lot",
        entity_id=lot_id,
        project_id=project_id,
        description=f"Excluiu o lote {old_data['code']} do projeto.",
        old_data=old_data,
        request=request,
        severity="CRITICAL",
    )


@router.get(
    "/lots/{lot_id}/link-candidates",
    response_model=list[LotLinkCandidateResponse],
)
def list_lot_link_candidates(
    project_id: UUID,
    lot_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[LotLinkCandidateResponse]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seals = (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id.is_(None),
        )
        .order_by(Seal.seal_code.asc())
        .all()
    )

    result: list[LotLinkCandidateResponse] = []

    for seal in seals:
        social = (
            db.query(SocialRegistration)
            .filter(
                SocialRegistration.project_id == project_id,
                SocialRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

        physical = (
            db.query(PhysicalRegistration)
            .filter(
                PhysicalRegistration.project_id == project_id,
                PhysicalRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

        documents_count = (
            db.query(func.count(Document.id))
            .filter(
                Document.project_id == project_id,
                Document.seal_code == seal.seal_code,
            )
            .scalar()
            or 0
        )

        result.append(
            LotLinkCandidateResponse(
                seal_id=str(seal.id),
                seal_code=seal.seal_code,
                lot_code=seal.lot_code,
                responsible_name=social.responsible_name if social else None,
                latitude=seal.latitude,
                longitude=seal.longitude,
                geo_link_status=seal.geo_link_status,
                distance_m=None,
                has_social=social is not None,
                has_physical=physical is not None,
                documents_count=documents_count,
            )
        )

    return result


@router.patch(
    "/lots/{lot_id}/link-seal",
    response_model=LotLinkSealResponse,
)
def link_seal_to_lot(
    project_id: UUID,
    lot_id: UUID,
    payload: LotLinkSealRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> LotLinkSealResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seal = (
        db.query(Seal)
        .filter(
            Seal.id == UUID(payload.seal_id),
            Seal.project_id == project_id,
        )
        .first()
    )

    if seal is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Selagem não encontrada.",
        )

    old_data = {
        "seal_id": str(seal.id),
        "seal_code": seal.seal_code,
        "old_lot_id": str(seal.lot_id) if seal.lot_id else None,
        "old_lot_code": seal.lot_code,
        "old_geo_link_status": seal.geo_link_status,
    }

    seal.lot_id = lot.id
    seal.lot_code = lot.code
    seal.geo_link_status = "confirmado"
    seal.geospatial_note = (
        f"Selagem vinculada manualmente ao lote {lot.code} pelo administrador."
    )

    lot.needs_review = True

    if lot.technical_status == "sem_geometria":
        lot.technical_status = "pendente"

    db.commit()
    db.refresh(seal)

    register_audit_log(
        db,
        user=current_user,
        action="UPDATE",
        entity_type="lot_link",
        entity_id=lot.id,
        project_id=project_id,
        description=f"Vinculou a selagem {seal.seal_code} ao lote {lot.code}.",
        old_data=old_data,
        new_data={
            "seal_id": str(seal.id),
            "seal_code": seal.seal_code,
            "lot_id": str(lot.id),
            "lot_code": lot.code,
            "geo_link_status": seal.geo_link_status,
        },
        request=request,
        severity="WARNING",
    )

    return LotLinkSealResponse(
        lot_id=str(lot.id),
        lot_code=lot.code,
        seal_id=str(seal.id),
        seal_code=seal.seal_code,
        message=f"Selagem {seal.seal_code} vinculada ao lote {lot.code}.",
    )


@router.patch("/lots/{lot_id}/unlink-seal", response_model=LotLinkSealResponse)
def unlink_seal_from_lot(
    project_id: UUID,
    lot_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> LotLinkSealResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seal = (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
        )
        .first()
    )

    if seal is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Este lote não possui selagem vinculada.",
        )

    old_data = {
        "seal_id": str(seal.id),
        "seal_code": seal.seal_code,
        "lot_id": str(lot.id),
        "lot_code": lot.code,
        "geo_link_status": seal.geo_link_status,
    }

    seal.lot_id = None
    seal.geo_link_status = "nao_validado"
    seal.geospatial_note = (
        f"Selagem desvinculada manualmente do lote {lot.code} pelo administrador."
    )

    lot.needs_review = True
    lot.is_ready_for_technical_documents = False

    if lot.lot_review_status == "apto":
        lot.lot_review_status = "em_revisao"
        lot.technical_status = "em_revisao"

    db.commit()
    db.refresh(seal)

    register_audit_log(
        db,
        user=current_user,
        action="UPDATE",
        entity_type="lot_unlink",
        entity_id=lot.id,
        project_id=project_id,
        description=f"Desvinculou a selagem {old_data['seal_code']} do lote {lot.code}.",
        old_data=old_data,
        new_data={
            "seal_id": str(seal.id),
            "seal_code": seal.seal_code,
            "lot_id": None,
            "geo_link_status": seal.geo_link_status,
        },
        request=request,
        severity="WARNING",
    )

    return LotLinkSealResponse(
        lot_id=str(lot.id),
        lot_code=lot.code,
        seal_id=str(seal.id),
        seal_code=seal.seal_code,
        message=f"Selagem {seal.seal_code} desvinculada do lote {lot.code}.",
    )


@router.get("/lots/{lot_id}/detail", response_model=ProjectMapLotResponse)
def get_lot_detail(
    project_id: UUID,
    lot_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ProjectMapLotResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seal = (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
        )
        .order_by(Seal.created_at.asc())
        .first()
    )

    social = None
    physical = None
    documents_count = 0

    if seal is not None:
        social = (
            db.query(SocialRegistration)
            .filter(
                SocialRegistration.project_id == project_id,
                SocialRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

        physical = (
            db.query(PhysicalRegistration)
            .filter(
                PhysicalRegistration.project_id == project_id,
                PhysicalRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

        documents_count = (
            db.query(func.count(Document.id))
            .filter(
                Document.project_id == project_id,
                Document.seal_code == seal.seal_code,
            )
            .scalar()
            or 0
        )

    pending_flags = _build_lot_pending_flags(
        lot=lot,
        seal=seal,
        social=social,
        physical=physical,
        documents_count=documents_count,
    )

    return ProjectMapLotResponse(
        id=str(lot.id),
        code=lot.code,
        block=lot.block,
        area_m2=lot.area_m2,
        perimeter_m=lot.perimeter_m,
        status=lot.status,
        needs_review=lot.needs_review,
        lot_review_status=getattr(lot, "lot_review_status", "preliminar"),
        technical_status=getattr(lot, "technical_status", "sem_geometria"),
        is_ready_for_technical_documents=getattr(
            lot,
            "is_ready_for_technical_documents",
            False,
        ),
        geometry_geojson=getattr(lot, "geometry_geojson", None),
        centroid_latitude=getattr(lot, "centroid_latitude", None),
        centroid_longitude=getattr(lot, "centroid_longitude", None),
        geospatial_source=getattr(lot, "geospatial_source", None),
        geospatial_accuracy_m=getattr(lot, "geospatial_accuracy_m", None),
        revision_notes=getattr(lot, "revision_notes", None),
        seal=(
            ProjectMapSealResponse(
                id=str(seal.id),
                seal_code=seal.seal_code,
                lot_code=seal.lot_code,
                situation=seal.situation,
                geo_link_status=seal.geo_link_status,
                needs_rtk_validation=seal.needs_rtk_validation,
                geospatial_note=seal.geospatial_note,
                latitude=seal.latitude,
                longitude=seal.longitude,
                gps_accuracy=seal.gps_accuracy,
            )
            if seal
            else None
        ),
        social=(
            ProjectMapSocialResponse(
                id=str(social.id),
                responsible_name=social.responsible_name,
                responsible_cpf=social.responsible_cpf,
                phone=social.phone,
                household_members=social.household_members,
                family_income=social.family_income,
                has_conflict=social.has_conflict,
            )
            if social
            else None
        ),
        physical=(
            ProjectMapPhysicalResponse(
                id=str(physical.id),
                property_type=physical.property_type,
                property_use=physical.property_use,
                wall_material=physical.wall_material,
                roof_type=physical.roof_type,
                floor_type=physical.floor_type,
                rooms=physical.rooms,
                bathrooms=physical.bathrooms,
                has_energy=physical.has_energy,
                has_water=physical.has_water,
                has_sewage=physical.has_sewage,
                has_bathroom=physical.has_bathroom,
                habitability_condition=physical.habitability_condition,
                risk_area=physical.risk_area,
                flood_prone=physical.flood_prone,
            )
            if physical
            else None
        ),
        documents_count=documents_count,
        pending_flags=pending_flags,
    )


@router.get(
    "/lots/{lot_id}/documents",
    response_model=list[LotDocumentResponse],
)
def list_lot_documents(
    project_id: UUID,
    lot_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[LotDocumentResponse]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = _get_lot_or_404(db, project_id=project_id, lot_id=lot_id)

    documents = _query_lot_documents(
        db,
        project_id=project_id,
        lot=lot,
    )

    return [_lot_document_to_response(document) for document in documents]


DOCUMENT_STORAGE_DIR = Path("storage/documents")


@router.post("/lots/{lot_id}/documents/upload", response_model=DocumentResponse)
async def upload_lot_document(
    project_id: UUID,
    lot_id: UUID,
    file: UploadFile = File(...),
    document_type: str = Form(...),
    notes: str | None = Form(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DocumentResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seal = (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
        )
        .first()
    )

    social = None

    if seal is not None:
        social = (
            db.query(SocialRegistration)
            .filter(
                SocialRegistration.project_id == project_id,
                SocialRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

    original_filename = file.filename or "documento"
    suffix = Path(original_filename).suffix.lower()

    if not suffix:
        guessed = mimetypes.guess_extension(file.content_type or "")
        suffix = guessed or ".bin"

    stored_filename = f"{uuid.uuid4()}{suffix}"

    target_dir = DOCUMENT_STORAGE_DIR / str(project_id) / str(lot_id)
    target_dir.mkdir(parents=True, exist_ok=True)

    target_path = target_dir / stored_filename

    content = await file.read()
    target_path.write_bytes(content)

    document = Document(
        project_id=project_id,
        lot_id=lot.id,
        seal_id=seal.id if seal else None,
        social_registration_id=social.id if social else None,
        seal_code=seal.seal_code if seal else f"LOTE-{lot.code}",
        document_type=document_type,
        file_path=str(target_path),
        original_filename=original_filename,
        stored_filename=stored_filename,
        mime_type=file.content_type or mimetypes.guess_type(original_filename)[0],
        file_size_bytes=len(content),
        notes=notes,
        validated=False,
        document_status="pendente",
    )

    db.add(document)
    db.commit()
    db.refresh(document)

    return _document_to_response(document)


@router.patch("/documents/{document_id}/validate", response_model=DocumentResponse)
def validate_project_document(
    project_id: UUID,
    document_id: UUID,
    payload: DocumentValidateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DocumentResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    document = (
        db.query(Document)
        .filter(
            Document.id == document_id,
            Document.project_id == project_id,
        )
        .first()
    )

    if document is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Documento não encontrado.",
        )

    old_data = {
        "validated": document.validated,
    }

    document.validated = payload.validated

    db.commit()
    db.refresh(document)

    register_audit_log(
        db,
        user=current_user,
        action="UPDATE",
        entity_type="document",
        entity_id=document.id,
        project_id=project_id,
        description=(
            "Validou documento do projeto."
            if document.validated
            else "Removeu validação de documento do projeto."
        ),
        old_data=old_data,
        new_data={
            "validated": document.validated,
        },
        request=request,
        severity="INFO",
    )

    return _document_to_response(document)


@router.delete(
    "/documents/{document_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
def delete_project_document(
    project_id: UUID,
    document_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> None:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    document = (
        db.query(Document)
        .filter(
            Document.id == document_id,
            Document.project_id == project_id,
        )
        .first()
    )

    if document is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Documento não encontrado.",
        )

    old_data = {
        "id": str(document.id),
        "lot_id": str(document.lot_id) if getattr(document, "lot_id", None) else None,
        "seal_id": str(document.seal_id) if document.seal_id else None,
        "seal_code": document.seal_code,
        "document_type": document.document_type,
        "file_path": document.file_path,
        "validated": document.validated,
        "document_status": getattr(document, "document_status", None),
    }

    file_path = Path(document.file_path)

    db.delete(document)
    db.commit()

    if file_path.exists() and file_path.is_file():
        try:
            file_path.unlink()
        except Exception:
            pass

    register_audit_log(
        db,
        user=current_user,
        action="DELETE",
        entity_type="document",
        entity_id=document_id,
        project_id=project_id,
        description=f"Excluiu documento {old_data['document_type']} do projeto.",
        old_data=old_data,
        request=request,
        severity="CRITICAL",
    )

    return None


@router.get("/documents/{document_id}/file")
def get_project_document_file(
    project_id: UUID,
    document_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    document = (
        db.query(Document)
        .filter(
            Document.id == document_id,
            Document.project_id == project_id,
        )
        .first()
    )

    if document is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Documento não encontrado.",
        )

    file_path = _resolve_document_path(document)

    if file_path is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                "O registro do documento existe, mas o arquivo físico não foi localizado. "
                "Verifique se o arquivo foi copiado do pacote mobile para o storage do backend."
            ),
        )

    # Se veio do mobile/imports/extracted, copia para storage/documents e atualiza o registro.
    if "storage/imports" in str(file_path).replace("\\", "/"):
        new_path = _copy_mobile_document_to_project_storage(
            project_id=project_id,
            source_path=file_path,
            fallback_filename=file_path.name,
        )

        document.file_path = new_path
        db.commit()
        db.refresh(document)

        file_path = Path(new_path)

    filename = file_path.name
    media_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type=media_type,
    )


@router.patch("/seals/{seal_id}", response_model=SealResponse)
def update_project_seal(
    project_id: UUID,
    seal_id: UUID,
    payload: SealUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> SealResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    seal = (
        db.query(Seal)
        .filter(
            Seal.id == seal_id,
            Seal.project_id == project_id,
        )
        .first()
    )

    if seal is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Selagem não encontrada.",
        )

    if payload.seal_code is not None:
        seal.seal_code = payload.seal_code.strip() or seal.seal_code

    if payload.lot_code is not None:
        seal.lot_code = payload.lot_code.strip() or None

    if payload.situation is not None:
        seal.situation = payload.situation.strip() or seal.situation

    if payload.geo_link_status is not None:
        seal.geo_link_status = payload.geo_link_status.strip() or seal.geo_link_status

    if payload.needs_rtk_validation is not None:
        seal.needs_rtk_validation = payload.needs_rtk_validation

    if payload.geospatial_note is not None:
        seal.geospatial_note = payload.geospatial_note.strip() or None

    if payload.latitude is not None:
        seal.latitude = payload.latitude

    if payload.longitude is not None:
        seal.longitude = payload.longitude

    if payload.gps_accuracy is not None:
        seal.gps_accuracy = payload.gps_accuracy

    db.commit()
    db.refresh(seal)

    return _seal_to_response(seal)


@router.patch(
    "/social-registrations/{social_id}", response_model=SocialRegistrationResponse
)
def update_project_social_registration(
    project_id: UUID,
    social_id: UUID,
    payload: SocialRegistrationUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> SocialRegistrationResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    social = (
        db.query(SocialRegistration)
        .filter(
            SocialRegistration.id == social_id,
            SocialRegistration.project_id == project_id,
        )
        .first()
    )

    if social is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cadastro social não encontrado.",
        )

    if payload.seal_code is not None:
        social.seal_code = payload.seal_code.strip() or social.seal_code

    if payload.responsible_name is not None:
        social.responsible_name = (
            payload.responsible_name.strip() or social.responsible_name
        )

    if payload.responsible_cpf is not None:
        social.responsible_cpf = payload.responsible_cpf.strip() or None

    if payload.responsible_rg is not None:
        social.responsible_rg = payload.responsible_rg.strip() or None

    if payload.issuing_agency is not None:
        social.issuing_agency = payload.issuing_agency.strip() or None

    if payload.phone is not None:
        social.phone = payload.phone.strip() or None

    if payload.marital_status is not None:
        social.marital_status = payload.marital_status.strip() or None

    if payload.profession is not None:
        social.profession = payload.profession.strip() or None

    if payload.household_members is not None:
        social.household_members = payload.household_members

    if payload.family_income is not None:
        social.family_income = payload.family_income

    if payload.receives_social_program is not None:
        social.receives_social_program = payload.receives_social_program

    if payload.social_program is not None:
        social.social_program = payload.social_program.strip() or None

    if payload.occupation_years is not None:
        social.occupation_years = payload.occupation_years

    if payload.occupation_type is not None:
        social.occupation_type = payload.occupation_type.strip() or None

    if payload.possession_document is not None:
        social.possession_document = payload.possession_document.strip() or None

    if payload.owns_other_property is not None:
        social.owns_other_property = payload.owns_other_property

    if payload.has_conflict is not None:
        social.has_conflict = payload.has_conflict

    if payload.notes is not None:
        social.notes = payload.notes.strip() or None

    db.commit()
    db.refresh(social)

    return _social_to_response(social)


@router.delete("/social-registrations/{social_id}")
def delete_project_social_registration(
    project_id: UUID,
    social_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, str]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    social = (
        db.query(SocialRegistration)
        .filter(
            SocialRegistration.id == social_id,
            SocialRegistration.project_id == project_id,
        )
        .first()
    )

    if social is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cadastro social não encontrado.",
        )

    documents_count = (
        db.query(Document)
        .filter(
            Document.project_id == project_id,
            Document.social_registration_id == social.id,
        )
        .count()
    )

    if documents_count > 0:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Este cadastro social possui documentos vinculados. "
                "Exclua ou desvincule os documentos antes de remover o cadastro."
            ),
        )

    db.delete(social)
    db.commit()

    return {"message": "Cadastro social excluído com sucesso."}


@router.patch(
    "/physical-registrations/{physical_id}", response_model=PhysicalRegistrationResponse
)
def update_project_physical_registration(
    project_id: UUID,
    physical_id: UUID,
    payload: PhysicalRegistrationUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> PhysicalRegistrationResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    physical = (
        db.query(PhysicalRegistration)
        .filter(
            PhysicalRegistration.id == physical_id,
            PhysicalRegistration.project_id == project_id,
        )
        .first()
    )

    if physical is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cadastro físico não encontrado.",
        )

    if payload.seal_code is not None:
        physical.seal_code = payload.seal_code.strip() or physical.seal_code

    if payload.property_type is not None:
        physical.property_type = payload.property_type.strip() or None

    if payload.property_use is not None:
        physical.property_use = payload.property_use.strip() or None

    if payload.wall_material is not None:
        physical.wall_material = payload.wall_material.strip() or None

    if payload.roof_type is not None:
        physical.roof_type = payload.roof_type.strip() or None

    if payload.floor_type is not None:
        physical.floor_type = payload.floor_type.strip() or None

    if payload.floors is not None:
        physical.floors = payload.floors

    if payload.rooms is not None:
        physical.rooms = payload.rooms

    if payload.bathrooms is not None:
        physical.bathrooms = payload.bathrooms

    if payload.has_energy is not None:
        physical.has_energy = payload.has_energy

    if payload.has_water is not None:
        physical.has_water = payload.has_water

    if payload.has_sewage is not None:
        physical.has_sewage = payload.has_sewage

    if payload.has_bathroom is not None:
        physical.has_bathroom = payload.has_bathroom

    if payload.habitability_condition is not None:
        physical.habitability_condition = payload.habitability_condition.strip() or None

    if payload.risk_area is not None:
        physical.risk_area = payload.risk_area

    if payload.flood_prone is not None:
        physical.flood_prone = payload.flood_prone

    if payload.notes is not None:
        physical.notes = payload.notes.strip() or None

    db.commit()
    db.refresh(physical)

    return _physical_to_response(physical)


@router.delete("/physical-registrations/{physical_id}")
def delete_project_physical_registration(
    project_id: UUID,
    physical_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, str]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    physical = (
        db.query(PhysicalRegistration)
        .filter(
            PhysicalRegistration.id == physical_id,
            PhysicalRegistration.project_id == project_id,
        )
        .first()
    )

    if physical is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cadastro físico não encontrado.",
        )

    db.delete(physical)
    db.commit()

    return {"message": "Cadastro físico excluído com sucesso."}


@router.get("/exports/summary")
def get_project_export_summary(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    project = db.query(Project).filter(Project.id == project_id).first()

    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projeto não encontrado.",
        )

    total_lots = db.query(Lot).filter(Lot.project_id == project_id).count()

    lots_with_geometry = (
        db.query(Lot)
        .filter(
            Lot.project_id == project_id,
            Lot.geometry_geojson.isnot(None),
        )
        .count()
    )

    ready_lots = (
        db.query(Lot)
        .filter(
            Lot.project_id == project_id,
            Lot.is_ready_for_technical_documents == True,
        )
        .count()
    )

    pending_lots = max(total_lots - ready_lots, 0)

    total_seals = db.query(Seal).filter(Seal.project_id == project_id).count()

    total_social = (
        db.query(SocialRegistration)
        .filter(SocialRegistration.project_id == project_id)
        .count()
    )

    total_physical = (
        db.query(PhysicalRegistration)
        .filter(PhysicalRegistration.project_id == project_id)
        .count()
    )

    total_documents = (
        db.query(Document).filter(Document.project_id == project_id).count()
    )

    validated_documents = (
        db.query(Document)
        .filter(
            Document.project_id == project_id,
            Document.validated == True,
        )
        .count()
    )

    return {
        "project_id": str(project.id),
        "project_name": project.name,
        "municipality": project.municipality,
        "state": project.state,
        "neighborhood": project.neighborhood,
        "total_lots": total_lots,
        "lots_with_geometry": lots_with_geometry,
        "ready_lots": ready_lots,
        "pending_lots": pending_lots,
        "total_seals": total_seals,
        "total_social_registrations": total_social,
        "total_physical_registrations": total_physical,
        "total_documents": total_documents,
        "validated_documents": validated_documents,
        "can_export_metricatopo": ready_lots > 0,
        "generated_at": datetime.now().isoformat(),
    }


@router.get("/exports/matrix")
def export_project_matrix(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> FileResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    project = db.query(Project).filter(Project.id == project_id).first()

    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projeto não encontrado.",
        )

    lots = (
        db.query(Lot)
        .filter(Lot.project_id == project_id)
        .order_by(Lot.code.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz REURB"

    title = f"MATRIZ TÉCNICO-CADASTRAL REURB - {project.name}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=31)
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="166534")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=31)
    ws.cell(row=2, column=1).value = (
        f"Município: {_safe_export_text(project.municipality)} / "
        f"{_safe_export_text(project.state)} | "
        f"Bairro/Núcleo: {_safe_export_text(project.neighborhood)} | "
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws.cell(row=2, column=1).font = Font(bold=True, color="334155")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = [
        "Código do lote",
        "Quadra",
        "Área m²",
        "Perímetro m",
        "Status do lote",
        "Apto peças técnicas",
        "Código da selagem",
        "Situação da selagem",
        "Vínculo geográfico",
        "Necessita RTK",
        "Latitude",
        "Longitude",
        "Responsável",
        "CPF",
        "RG",
        "Telefone",
        "Estado civil",
        "Profissão",
        "Moradores",
        "Renda familiar",
        "Programa social",
        "Tempo ocupação anos",
        "Forma ocupação",
        "Documento posse",
        "Possui outro imóvel",
        "Conflito",
        "Tipo imóvel",
        "Uso imóvel",
        "Cômodos",
        "Banheiros",
        "Documentos",
        "Documentos validados",
        "Pendências / observações",
    ]

    header_row = 4

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    row_number = header_row + 1

    for lot in lots:
        seal = (
            db.query(Seal)
            .filter(
                Seal.project_id == project_id,
                Seal.lot_id == lot.id,
            )
            .first()
        )

        if seal is None:
            seal = (
                db.query(Seal)
                .filter(
                    Seal.project_id == project_id,
                    Seal.lot_code == lot.code,
                )
                .first()
            )

        social = None
        physical = None

        if seal:
            social = (
                db.query(SocialRegistration)
                .filter(
                    SocialRegistration.project_id == project_id,
                    SocialRegistration.seal_code == seal.seal_code,
                )
                .first()
            )

            physical = (
                db.query(PhysicalRegistration)
                .filter(
                    PhysicalRegistration.project_id == project_id,
                    PhysicalRegistration.seal_code == seal.seal_code,
                )
                .first()
            )

        documents_query = db.query(Document).filter(Document.project_id == project_id)

        if seal:
            documents_query = documents_query.filter(
                or_(
                    Document.lot_id == lot.id,
                    Document.seal_id == seal.id,
                    Document.seal_code == seal.seal_code,
                )
            )
        else:
            documents_query = documents_query.filter(Document.lot_id == lot.id)

        documents_count = documents_query.count()
        validated_documents_count = documents_query.filter(
            Document.validated == True
        ).count()

        pending_notes: list[str] = []

        if not lot.geometry_geojson:
            pending_notes.append("Sem geometria")

        if not seal:
            pending_notes.append("Sem selagem vinculada")

        if seal and seal.needs_rtk_validation:
            pending_notes.append("Necessita validação RTK")

        if not social:
            pending_notes.append("Sem cadastro social")

        if social and social.has_conflict:
            pending_notes.append("Possui conflito social")

        if not physical:
            pending_notes.append("Sem cadastro físico")

        if physical and physical.risk_area:
            pending_notes.append("Área de risco")

        if physical and physical.flood_prone:
            pending_notes.append("Sujeito à inundação")

        if documents_count == 0:
            pending_notes.append("Sem documentos")

        if documents_count > 0 and validated_documents_count < documents_count:
            pending_notes.append("Documento pendente de validação")

        row = [
            lot.code,
            lot.block,
            lot.area_m2,
            lot.perimeter_m,
            _status_apto(lot),
            _yes_no(getattr(lot, "is_ready_for_technical_documents", False)),
            seal.seal_code if seal else "",
            seal.situation if seal else "",
            seal.geo_link_status if seal else "",
            _yes_no(seal.needs_rtk_validation) if seal else "",
            seal.latitude if seal else "",
            seal.longitude if seal else "",
            social.responsible_name if social else "",
            social.responsible_cpf if social else "",
            social.responsible_rg if social else "",
            social.phone if social else "",
            social.marital_status if social else "",
            social.profession if social else "",
            social.household_members if social else "",
            social.family_income if social else "",
            (
                social.social_program
                if social and social.receives_social_program
                else "Não"
            ),
            social.occupation_years if social else "",
            social.occupation_type if social else "",
            social.possession_document if social else "",
            _yes_no(social.owns_other_property) if social else "",
            _yes_no(social.has_conflict) if social else "",
            physical.property_type if physical else "",
            physical.property_use if physical else "",
            physical.rooms if physical else "",
            physical.bathrooms if physical else "",
            documents_count,
            validated_documents_count,
            "; ".join(pending_notes) if pending_notes else "Sem pendências",
        ]

        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=row_number, column=col)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        row_number += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{row_number - 1}"

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 42)

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    project_slug = _safe_filename_slug(project.name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    filename = f"matriz_reurb_{project_slug}_{timestamp}.xlsx"
    output_path = EXPORTS_DIR / str(project_id)
    output_path.mkdir(parents=True, exist_ok=True)

    file_path = output_path / filename
    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/exports/geospatial-package")
def export_project_geospatial_package(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> FileResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    project = db.query(Project).filter(Project.id == project_id).first()

    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projeto não encontrado.",
        )

    lots = (
        db.query(Lot)
        .filter(Lot.project_id == project_id)
        .order_by(Lot.code.asc())
        .all()
    )

    seals = (
        db.query(Seal)
        .filter(Seal.project_id == project_id)
        .order_by(Seal.seal_code.asc())
        .all()
    )

    def lot_feature(lot: Lot) -> dict | None:
        geometry = _load_lot_geojson_geometry(lot.geometry_geojson)

        if geometry is None:
            return None

        seal = (
            db.query(Seal)
            .filter(
                Seal.project_id == project_id,
                Seal.lot_id == lot.id,
            )
            .first()
        )

        social = None
        physical = None

        if seal:
            social = (
                db.query(SocialRegistration)
                .filter(
                    SocialRegistration.project_id == project_id,
                    SocialRegistration.seal_code == seal.seal_code,
                )
                .first()
            )

            physical = (
                db.query(PhysicalRegistration)
                .filter(
                    PhysicalRegistration.project_id == project_id,
                    PhysicalRegistration.seal_code == seal.seal_code,
                )
                .first()
            )

        documents_query = db.query(Document).filter(Document.project_id == project_id)

        if seal:
            documents_query = documents_query.filter(
                or_(
                    Document.lot_id == lot.id,
                    Document.seal_id == seal.id,
                    Document.seal_code == seal.seal_code,
                )
            )
        else:
            documents_query = documents_query.filter(Document.lot_id == lot.id)

        documents_count = documents_query.count()
        validated_documents_count = documents_query.filter(
            Document.validated == True
        ).count()

        return {
            "type": "Feature",
            "geometry": geometry,
            "properties": {
                "lot_id": str(lot.id),
                "codigo_lote": lot.code,
                "quadra": lot.block,
                "area_m2": lot.area_m2,
                "perimetro_m": lot.perimeter_m,
                "status_lote": lot.status,
                "status_tecnico": getattr(lot, "technical_status", None),
                "apto_pecas": bool(
                    getattr(lot, "is_ready_for_technical_documents", False)
                ),
                "codigo_selo": seal.seal_code if seal else None,
                "situacao_selagem": seal.situation if seal else None,
                "vinculo_geo": seal.geo_link_status if seal else None,
                "necessita_rtk": seal.needs_rtk_validation if seal else None,
                "responsavel": social.responsible_name if social else None,
                "cpf": social.responsible_cpf if social else None,
                "telefone": social.phone if social else None,
                "tipo_imovel": physical.property_type if physical else None,
                "uso_imovel": physical.property_use if physical else None,
                "comodos": physical.rooms if physical else None,
                "banheiros": physical.bathrooms if physical else None,
                "documentos": documents_count,
                "documentos_validados": validated_documents_count,
                "fonte": lot.source_file,
            },
        }

    def seal_feature(seal: Seal) -> dict | None:
        if seal.latitude is None or seal.longitude is None:
            return None

        social = (
            db.query(SocialRegistration)
            .filter(
                SocialRegistration.project_id == project_id,
                SocialRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

        return {
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [seal.longitude, seal.latitude],
            },
            "properties": {
                "seal_id": str(seal.id),
                "codigo_selo": seal.seal_code,
                "lot_id": str(seal.lot_id) if seal.lot_id else None,
                "codigo_lote": seal.lot_code,
                "situacao": seal.situation,
                "vinculo_geo": seal.geo_link_status,
                "necessita_rtk": seal.needs_rtk_validation,
                "precisao_gps": seal.gps_accuracy,
                "responsavel": social.responsible_name if social else None,
                "cpf": social.responsible_cpf if social else None,
                "telefone": social.phone if social else None,
            },
        }

    lot_features = [
        feature for feature in (lot_feature(lot) for lot in lots) if feature
    ]
    seal_features = [
        feature for feature in (seal_feature(seal) for seal in seals) if feature
    ]

    if not lot_features:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Nenhuma geometria válida de lote foi encontrada para exportação. "
                "Verifique se os lotes possuem geometry_geojson preenchido corretamente."
            ),
        )

    lots_all_geojson = {
        "type": "FeatureCollection",
        "name": "lotes_todos",
        "features": lot_features,
    }

    lots_ready_geojson = {
        "type": "FeatureCollection",
        "name": "lotes_aptos",
        "features": [
            feature
            for feature in lot_features
            if feature["properties"].get("apto_pecas") is True
        ],
    }

    lots_pending_geojson = {
        "type": "FeatureCollection",
        "name": "lotes_pendentes",
        "features": [
            feature
            for feature in lot_features
            if feature["properties"].get("apto_pecas") is not True
        ],
    }

    seals_geojson = {
        "type": "FeatureCollection",
        "name": "selagens",
        "features": seal_features,
    }

    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer, delimiter=";")

    writer.writerow(
        [
            "codigo_lote",
            "quadra",
            "area_m2",
            "perimetro_m",
            "status_lote",
            "apto_pecas",
            "codigo_selo",
            "responsavel",
            "cpf",
            "telefone",
            "tipo_imovel",
            "uso_imovel",
            "documentos",
            "documentos_validados",
        ]
    )

    for feature in lot_features:
        props = feature["properties"]

        writer.writerow(
            [
                props.get("codigo_lote"),
                props.get("quadra"),
                props.get("area_m2"),
                props.get("perimetro_m"),
                props.get("status_lote"),
                "Sim" if props.get("apto_pecas") else "Não",
                props.get("codigo_selo"),
                props.get("responsavel"),
                props.get("cpf"),
                props.get("telefone"),
                props.get("tipo_imovel"),
                props.get("uso_imovel"),
                props.get("documentos"),
                props.get("documentos_validados"),
            ]
        )

    metadata = f"""PACOTE GEOESPACIAL BIOME REURB

Projeto: {project.name}
Município/UF: {project.municipality}/{project.state}
Bairro/Núcleo: {project.neighborhood}
Data de geração: {datetime.now().strftime("%d/%m/%Y %H:%M")}

Sistema de referência:
- Coordenadas geográficas em SIRGAS 2000 / EPSG:4326, quando aplicável.

Conteúdo:
01_lotes_todos.geojson
02_lotes_aptos.geojson
03_lotes_pendentes.geojson
04_selagens.geojson
05_atributos_lotes.csv
06_metadados.txt
kml/lotes_todos.kml
kml/lotes_aptos.kml
kml/lotes_pendentes.kml
kml/selagens.kml
shp/lotes_todos.zip
shp/lotes_aptos.zip
shp/lotes_pendentes.zip
shp/selagens.zip

Observação:
Este pacote é produto técnico-cadastral e geoespacial consolidado pelo BIOME REURB.
As plantas e memoriais descritivos formais poderão ser gerados posteriormente em software topográfico, como Métrica TOPO, a partir da base saneada.
"""

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    project_slug = _safe_filename_slug(project.name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_dir = EXPORTS_DIR / str(project_id)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"pacote_geoespacial_reurb_{project_slug}_{timestamp}.zip"
    zip_path = output_dir / filename

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr(
            "01_lotes_todos.geojson",
            json.dumps(lots_all_geojson, ensure_ascii=False, indent=2),
        )
        zip_file.writestr(
            "02_lotes_aptos.geojson",
            json.dumps(lots_ready_geojson, ensure_ascii=False, indent=2),
        )
        zip_file.writestr(
            "03_lotes_pendentes.geojson",
            json.dumps(lots_pending_geojson, ensure_ascii=False, indent=2),
        )
        zip_file.writestr(
            "04_selagens.geojson",
            json.dumps(seals_geojson, ensure_ascii=False, indent=2),
        )
        zip_file.writestr(
            "05_atributos_lotes.csv",
            csv_buffer.getvalue(),
        )
        zip_file.writestr(
            "06_metadados.txt",
            metadata,
        )
        zip_file.writestr(
            "kml/lotes_todos.kml",
            _feature_collection_to_kml(lots_all_geojson, name="Lotes todos"),
        )

        zip_file.writestr(
            "kml/lotes_aptos.kml",
            _feature_collection_to_kml(lots_ready_geojson, name="Lotes aptos"),
        )

        zip_file.writestr(
            "kml/lotes_pendentes.kml",
            _feature_collection_to_kml(lots_pending_geojson, name="Lotes pendentes"),
        )

        zip_file.writestr(
            "kml/selagens.kml",
            _feature_collection_to_kml(seals_geojson, name="Selagens"),
        )

        zip_file.writestr(
            "shp/lotes_todos.zip",
            _feature_collection_to_shapefile_zip(
                lots_all_geojson,
                name="lotes_todos",
                geometry_type="polygon",
            ),
        )

        zip_file.writestr(
            "shp/lotes_aptos.zip",
            _feature_collection_to_shapefile_zip(
                lots_ready_geojson,
                name="lotes_aptos",
                geometry_type="polygon",
            ),
        )

        zip_file.writestr(
            "shp/lotes_pendentes.zip",
            _feature_collection_to_shapefile_zip(
                lots_pending_geojson,
                name="lotes_pendentes",
                geometry_type="polygon",
            ),
        )

        zip_file.writestr(
            "shp/selagens.zip",
            _feature_collection_to_shapefile_zip(
                seals_geojson,
                name="selagens",
                geometry_type="point",
            ),
        )

    return FileResponse(
        path=zip_path,
        filename=filename,
        media_type="application/zip",
    )


@router.get("/exports/metricatopo-package")
def export_project_metricatopo_package(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> FileResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    project = db.query(Project).filter(Project.id == project_id).first()

    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projeto não encontrado.",
        )

    lots = (
        db.query(Lot)
        .filter(
            Lot.project_id == project_id,
            Lot.is_ready_for_technical_documents == True,
            Lot.geometry_geojson.isnot(None),
        )
        .order_by(Lot.code.asc())
        .all()
    )

    if not lots:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Não existem lotes aptos com geometria para exportação ao Métrica TOPO.",
        )

    def _coords_from_geometry(geometry: dict) -> list[list[float]]:
        """
        Extrai o primeiro anel externo de Polygon/MultiPolygon.
        Retorna coordenadas em [lon, lat].
        """
        geom_type = geometry.get("type")
        coordinates = geometry.get("coordinates")

        if not coordinates:
            return []

        if geom_type == "Polygon":
            if coordinates and coordinates[0]:
                return coordinates[0]

        if geom_type == "MultiPolygon":
            if coordinates and coordinates[0] and coordinates[0][0]:
                return coordinates[0][0]

        return []

    def _lot_context(lot: Lot):
        seal = (
            db.query(Seal)
            .filter(
                Seal.project_id == project_id,
                Seal.lot_id == lot.id,
            )
            .first()
        )

        if seal is None:
            seal = (
                db.query(Seal)
                .filter(
                    Seal.project_id == project_id,
                    Seal.lot_code == lot.code,
                )
                .first()
            )

        social = None
        physical = None

        if seal:
            social = (
                db.query(SocialRegistration)
                .filter(
                    SocialRegistration.project_id == project_id,
                    SocialRegistration.seal_code == seal.seal_code,
                )
                .first()
            )

            physical = (
                db.query(PhysicalRegistration)
                .filter(
                    PhysicalRegistration.project_id == project_id,
                    PhysicalRegistration.seal_code == seal.seal_code,
                )
                .first()
            )

        documents_query = db.query(Document).filter(Document.project_id == project_id)

        if seal:
            documents_query = documents_query.filter(
                or_(
                    Document.lot_id == lot.id,
                    Document.seal_id == seal.id,
                    Document.seal_code == seal.seal_code,
                )
            )
        else:
            documents_query = documents_query.filter(Document.lot_id == lot.id)

        documents_count = documents_query.count()
        validated_documents_count = documents_query.filter(
            Document.validated == True
        ).count()

        return seal, social, physical, documents_count, validated_documents_count

    lot_features = []
    vertices_rows = []
    atributos_rows = []
    confrontantes_rows = []

    for lot in lots:
        geometry = _load_lot_geojson_geometry(lot.geometry_geojson)

        if geometry is None:
            continue

        seal, social, physical, documents_count, validated_documents_count = (
            _lot_context(lot)
        )

        properties = {
            "codigo_lote": lot.code,
            "quadra": lot.block,
            "area_m2": lot.area_m2,
            "perimetro_m": lot.perimeter_m,
            "status_tecnico": getattr(lot, "technical_status", None),
            "apto_pecas": bool(getattr(lot, "is_ready_for_technical_documents", False)),
            "codigo_selo": seal.seal_code if seal else None,
            "responsavel": social.responsible_name if social else None,
            "cpf": social.responsible_cpf if social else None,
            "telefone": social.phone if social else None,
            "tipo_imovel": physical.property_type if physical else None,
            "uso_imovel": physical.property_use if physical else None,
            "documentos": documents_count,
            "documentos_validados": validated_documents_count,
        }

        lot_features.append(
            {
                "type": "Feature",
                "geometry": geometry,
                "properties": properties,
            }
        )

        atributos_rows.append(properties)

        coords = _coords_from_geometry(geometry)

        # Remove o último vértice duplicado, quando o anel fecha repetindo o primeiro ponto.
        if len(coords) > 1 and coords[0] == coords[-1]:
            coords_to_export = coords[:-1]
        else:
            coords_to_export = coords

        for index, coord in enumerate(coords_to_export, start=1):
            if not isinstance(coord, list | tuple) or len(coord) < 2:
                continue

            lon = coord[0]
            lat = coord[1]

            vertices_rows.append(
                {
                    "codigo_lote": lot.code,
                    "vertice": f"{lot.code}_V{index:03d}",
                    "ordem": index,
                    "longitude": lon,
                    "latitude": lat,
                    "observacao": "Vértice extraído da geometria validada no BIOME REURB.",
                }
            )

        confrontantes_rows.append(
            {
                "codigo_lote": lot.code,
                "confrontante_norte": "",
                "confrontante_sul": "",
                "confrontante_leste": "",
                "confrontante_oeste": "",
                "observacao": "Preencher/validar confrontações no Métrica TOPO ou etapa topográfica.",
            }
        )

    if not lot_features:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Existem lotes aptos, mas nenhuma geometria válida foi encontrada "
                "para exportação ao Métrica TOPO."
            ),
        )

    lots_metricatopo_geojson = {
        "type": "FeatureCollection",
        "name": "lotes_aptos_metricatopo",
        "features": lot_features,
    }

    atributos_buffer = StringIO()
    atributos_writer = csv.writer(atributos_buffer, delimiter=";")
    atributos_writer.writerow(
        [
            "codigo_lote",
            "quadra",
            "area_m2",
            "perimetro_m",
            "status_tecnico",
            "apto_pecas",
            "codigo_selo",
            "responsavel",
            "cpf",
            "telefone",
            "tipo_imovel",
            "uso_imovel",
            "documentos",
            "documentos_validados",
        ]
    )

    for row in atributos_rows:
        atributos_writer.writerow(
            [
                row.get("codigo_lote"),
                row.get("quadra"),
                row.get("area_m2"),
                row.get("perimetro_m"),
                row.get("status_tecnico"),
                "Sim" if row.get("apto_pecas") else "Não",
                row.get("codigo_selo"),
                row.get("responsavel"),
                row.get("cpf"),
                row.get("telefone"),
                row.get("tipo_imovel"),
                row.get("uso_imovel"),
                row.get("documentos"),
                row.get("documentos_validados"),
            ]
        )

    vertices_buffer = StringIO()
    vertices_writer = csv.writer(vertices_buffer, delimiter=";")
    vertices_writer.writerow(
        [
            "codigo_lote",
            "vertice",
            "ordem",
            "longitude",
            "latitude",
            "observacao",
        ]
    )

    for row in vertices_rows:
        vertices_writer.writerow(
            [
                row["codigo_lote"],
                row["vertice"],
                row["ordem"],
                row["longitude"],
                row["latitude"],
                row["observacao"],
            ]
        )

    confrontantes_buffer = StringIO()
    confrontantes_writer = csv.writer(confrontantes_buffer, delimiter=";")
    confrontantes_writer.writerow(
        [
            "codigo_lote",
            "confrontante_norte",
            "confrontante_sul",
            "confrontante_leste",
            "confrontante_oeste",
            "observacao",
        ]
    )

    for row in confrontantes_rows:
        confrontantes_writer.writerow(
            [
                row["codigo_lote"],
                row["confrontante_norte"],
                row["confrontante_sul"],
                row["confrontante_leste"],
                row["confrontante_oeste"],
                row["observacao"],
            ]
        )

    metadata = f"""PACOTE MÉTRICA TOPO - BIOME REURB

Projeto: {project.name}
Município/UF: {project.municipality}/{project.state}
Bairro/Núcleo: {project.neighborhood}
Data de geração: {datetime.now().strftime("%d/%m/%Y %H:%M")}

Critério de seleção:
- Foram exportados apenas lotes marcados como aptos para peças técnicas.
- Foram exportados apenas lotes com geometria disponível no BIOME REURB.

Sistema de referência:
- Coordenadas geográficas em SIRGAS 2000 / EPSG:4326, quando aplicável.
- Os vértices foram extraídos da geometria validada armazenada no BIOME REURB.

Conteúdo:
01_lotes_aptos_metricatopo.geojson
02_atributos_lotes_aptos.csv
03_vertices_lotes_aptos.csv
04_confrontantes_preliminares.csv
05_metadados_metricatopo.txt
kml/lotes_aptos_metricatopo.kml
shp/lotes_aptos_metricatopo.zip

Observação técnica:
Este pacote não substitui a planta e o memorial descritivo formal.
Ele deve ser utilizado como base saneada para importação, conferência e geração das peças topográficas no Métrica TOPO.
As confrontações devem ser conferidas e ajustadas pela equipe técnica responsável.
"""

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    project_slug = _safe_filename_slug(project.name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_dir = EXPORTS_DIR / str(project_id)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"pacote_metricatopo_reurb_{project_slug}_{timestamp}.zip"
    zip_path = output_dir / filename

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr(
            "01_lotes_aptos_metricatopo.geojson",
            json.dumps(lots_metricatopo_geojson, ensure_ascii=False, indent=2),
        )
        zip_file.writestr(
            "02_atributos_lotes_aptos.csv",
            atributos_buffer.getvalue(),
        )
        zip_file.writestr(
            "03_vertices_lotes_aptos.csv",
            vertices_buffer.getvalue(),
        )
        zip_file.writestr(
            "04_confrontantes_preliminares.csv",
            confrontantes_buffer.getvalue(),
        )
        zip_file.writestr(
            "05_metadados_metricatopo.txt",
            metadata,
        )
        zip_file.writestr(
            "kml/lotes_aptos_metricatopo.kml",
            _feature_collection_to_kml(
                lots_metricatopo_geojson,
                name="Lotes aptos Métrica TOPO",
            ),
        )

        zip_file.writestr(
            "shp/lotes_aptos_metricatopo.zip",
            _feature_collection_to_shapefile_zip(
                lots_metricatopo_geojson,
                name="lotes_aptos_metricatopo",
                geometry_type="polygon",
            ),
        )

    return FileResponse(
        path=zip_path,
        filename=filename,
        media_type="application/zip",
    )


@router.get("/exports/lots-dossiers")
def export_project_lots_dossiers(
    project_id: UUID,
    only_ready: bool = Query(default=True),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> FileResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    project = db.query(Project).filter(Project.id == project_id).first()

    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projeto não encontrado.",
        )

    query = db.query(Lot).filter(Lot.project_id == project_id)

    if only_ready:
        query = query.filter(Lot.is_ready_for_technical_documents == True)

    lots = query.order_by(Lot.code.asc()).all()

    if not lots:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Nenhum lote encontrado para geração dos dossiês.",
        )

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    project_slug = _safe_filename_slug(project.name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_dir = EXPORTS_DIR / str(project_id)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"dossies_lotes_reurb_{project_slug}_{timestamp}.zip"
    zip_path = output_dir / filename

    indice_buffer = StringIO()
    indice_writer = csv.writer(indice_buffer, delimiter=";")
    indice_writer.writerow(
        [
            "codigo_lote",
            "apto_pecas",
            "codigo_selo",
            "responsavel",
            "cpf",
            "documentos",
            "documentos_validados",
            "pasta",
        ]
    )

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for lot in lots:
            folder_name = f"LOTE_{_safe_filename_slug(lot.code)}"

            seal = (
                db.query(Seal)
                .filter(
                    Seal.project_id == project_id,
                    Seal.lot_id == lot.id,
                )
                .first()
            )

            if seal is None:
                seal = (
                    db.query(Seal)
                    .filter(
                        Seal.project_id == project_id,
                        Seal.lot_code == lot.code,
                    )
                    .first()
                )

            social = None
            physical = None

            if seal:
                social = (
                    db.query(SocialRegistration)
                    .filter(
                        SocialRegistration.project_id == project_id,
                        SocialRegistration.seal_code == seal.seal_code,
                    )
                    .first()
                )

                physical = (
                    db.query(PhysicalRegistration)
                    .filter(
                        PhysicalRegistration.project_id == project_id,
                        PhysicalRegistration.seal_code == seal.seal_code,
                    )
                    .first()
                )

            documents_query = db.query(Document).filter(
                Document.project_id == project_id
            )

            if seal:
                documents_query = documents_query.filter(
                    or_(
                        Document.lot_id == lot.id,
                        Document.seal_id == seal.id,
                        Document.seal_code == seal.seal_code,
                    )
                )
            else:
                documents_query = documents_query.filter(Document.lot_id == lot.id)

            documents = documents_query.order_by(Document.document_type.asc()).all()
            documents_count = len(documents)
            validated_documents_count = len([doc for doc in documents if doc.validated])

            pdf_bytes = _generate_lot_dossier_pdf(
                project=project,
                lot=lot,
                seal=seal,
                social=social,
                physical=physical,
                documents_count=documents_count,
                validated_documents_count=validated_documents_count,
            )

            zip_file.writestr(
                f"{folder_name}/ficha_cadastral_{folder_name}.pdf",
                pdf_bytes,
            )

            geometry = _load_lot_geojson_geometry(lot.geometry_geojson)

            if geometry is not None:
                lot_geojson = {
                    "type": "FeatureCollection",
                    "name": folder_name,
                    "features": [
                        {
                            "type": "Feature",
                            "geometry": geometry,
                            "properties": {
                                "codigo_lote": lot.code,
                                "quadra": lot.block,
                                "area_m2": lot.area_m2,
                                "perimetro_m": lot.perimeter_m,
                                "apto_pecas": bool(
                                    getattr(
                                        lot,
                                        "is_ready_for_technical_documents",
                                        False,
                                    )
                                ),
                                "codigo_selo": seal.seal_code if seal else None,
                                "responsavel": (
                                    social.responsible_name if social else None
                                ),
                                "cpf": social.responsible_cpf if social else None,
                            },
                        }
                    ],
                }

                zip_file.writestr(
                    f"{folder_name}/geometria_{folder_name}.geojson",
                    json.dumps(lot_geojson, ensure_ascii=False, indent=2),
                )

                zip_file.writestr(
                    f"{folder_name}/geometria_{folder_name}.kml",
                    _feature_collection_to_kml(
                        lot_geojson,
                        name=f"Geometria {lot.code}",
                    ),
                )

                zip_file.writestr(
                    f"{folder_name}/shp/geometria_{folder_name}.zip",
                    _feature_collection_to_shapefile_zip(
                        lot_geojson,
                        name=f"geometria_{folder_name}"[:40],
                        geometry_type="polygon",
                    ),
                )

            atributos_buffer = StringIO()
            atributos_writer = csv.writer(atributos_buffer, delimiter=";")
            atributos_writer.writerow(["campo", "valor"])
            atributos_writer.writerow(["codigo_lote", lot.code])
            atributos_writer.writerow(["quadra", lot.block])
            atributos_writer.writerow(["area_m2", lot.area_m2])
            atributos_writer.writerow(["perimetro_m", lot.perimeter_m])
            atributos_writer.writerow(
                [
                    "apto_pecas",
                    _yes_no(getattr(lot, "is_ready_for_technical_documents", False)),
                ]
            )
            atributos_writer.writerow(["codigo_selo", seal.seal_code if seal else ""])
            atributos_writer.writerow(
                ["responsavel", social.responsible_name if social else ""]
            )
            atributos_writer.writerow(["cpf", social.responsible_cpf if social else ""])
            atributos_writer.writerow(["telefone", social.phone if social else ""])
            atributos_writer.writerow(
                ["tipo_imovel", physical.property_type if physical else ""]
            )
            atributos_writer.writerow(
                ["uso_imovel", physical.property_use if physical else ""]
            )
            atributos_writer.writerow(["documentos", documents_count])
            atributos_writer.writerow(
                ["documentos_validados", validated_documents_count]
            )

            zip_file.writestr(
                f"{folder_name}/atributos_{folder_name}.csv",
                atributos_buffer.getvalue(),
            )

            metadados = f"""DOSSIÊ INDIVIDUAL DO LOTE - BIOME REURB

Projeto: {project.name}
Município/UF: {project.municipality}/{project.state}
Bairro/Núcleo: {project.neighborhood}
Lote: {lot.code}
Data de geração: {datetime.now().strftime("%d/%m/%Y %H:%M")}

Conteúdo da pasta:
- Ficha cadastral PDF;
- Geometria individual em GeoJSON/KML/SHP, quando disponível;
- Atributos consolidados em CSV;
- Documentos vinculados ao lote/selagem, quando localizados no storage.

Observação:
Este dossiê subsidia a análise técnico-cadastral e documental do lote.
Não substitui a planta e o memorial descritivo formal.
"""

            zip_file.writestr(
                f"{folder_name}/metadados_{folder_name}.txt",
                metadados,
            )

            for document in documents:
                document_path = _resolve_export_document_path(document)

                if document_path is None:
                    continue

                original_name = (
                    document.original_filename
                    or document.stored_filename
                    or document_path.name
                )

                safe_doc_name = _safe_filename_slug(Path(original_name).stem)
                suffix = Path(original_name).suffix or document_path.suffix

                archive_name = (
                    f"{folder_name}/documentos/"
                    f"{safe_doc_name}_{str(document.id)[:8]}{suffix}"
                )

                zip_file.write(document_path, arcname=archive_name)

            indice_writer.writerow(
                [
                    lot.code,
                    _yes_no(getattr(lot, "is_ready_for_technical_documents", False)),
                    seal.seal_code if seal else "",
                    social.responsible_name if social else "",
                    social.responsible_cpf if social else "",
                    documents_count,
                    validated_documents_count,
                    folder_name,
                ]
            )

        zip_file.writestr("indice_dossies.csv", indice_buffer.getvalue())

    return FileResponse(
        path=zip_path,
        filename=filename,
        media_type="application/zip",
    )


@router.get("/lots/{lot_id}/dossier")
def export_single_lot_dossier(
    project_id: UUID,
    lot_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> FileResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    project = db.query(Project).filter(Project.id == project_id).first()

    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projeto não encontrado.",
        )

    lot = (
        db.query(Lot)
        .filter(
            Lot.id == lot_id,
            Lot.project_id == project_id,
        )
        .first()
    )

    if lot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lote não encontrado.",
        )

    seal = (
        db.query(Seal)
        .filter(
            Seal.project_id == project_id,
            Seal.lot_id == lot.id,
        )
        .first()
    )

    if seal is None:
        seal = (
            db.query(Seal)
            .filter(
                Seal.project_id == project_id,
                Seal.lot_code == lot.code,
            )
            .first()
        )

    social = None
    physical = None

    if seal:
        social = (
            db.query(SocialRegistration)
            .filter(
                SocialRegistration.project_id == project_id,
                SocialRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

        physical = (
            db.query(PhysicalRegistration)
            .filter(
                PhysicalRegistration.project_id == project_id,
                PhysicalRegistration.seal_code == seal.seal_code,
            )
            .first()
        )

    documents_query = db.query(Document).filter(Document.project_id == project_id)

    conditions = [
        Document.lot_id == lot.id,
    ]

    if seal:
        conditions.extend(
            [
                Document.seal_id == seal.id,
                Document.seal_code == seal.seal_code,
            ]
        )

    if social:
        conditions.append(Document.social_registration_id == social.id)

    documents = (
        documents_query.filter(or_(*conditions))
        .order_by(Document.document_type.asc())
        .all()
    )

    documents_count = len(documents)
    validated_documents_count = len([doc for doc in documents if doc.validated])

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    project_slug = _safe_filename_slug(project.name)
    lot_slug = _safe_filename_slug(lot.code)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_dir = EXPORTS_DIR / str(project_id)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"dossie_lote_{lot_slug}_{project_slug}_{timestamp}.zip"
    zip_path = output_dir / filename

    folder_name = f"LOTE_{lot_slug}"

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
        pdf_bytes = _generate_lot_dossier_pdf(
            project=project,
            lot=lot,
            seal=seal,
            social=social,
            physical=physical,
            documents_count=documents_count,
            validated_documents_count=validated_documents_count,
        )

        zip_file.writestr(
            f"{folder_name}/ficha_cadastral_{folder_name}.pdf",
            pdf_bytes,
        )

        geometry = _load_lot_geojson_geometry(lot.geometry_geojson)

        if geometry is not None:
            lot_geojson = {
                "type": "FeatureCollection",
                "name": folder_name,
                "features": [
                    {
                        "type": "Feature",
                        "geometry": geometry,
                        "properties": {
                            "codigo_lote": lot.code,
                            "quadra": lot.block,
                            "area_m2": lot.area_m2,
                            "perimetro_m": lot.perimeter_m,
                            "status": lot.status,
                            "status_tecnico": getattr(lot, "technical_status", None),
                            "apto_pecas": bool(
                                getattr(
                                    lot,
                                    "is_ready_for_technical_documents",
                                    False,
                                )
                            ),
                            "codigo_selo": seal.seal_code if seal else None,
                            "responsavel": social.responsible_name if social else None,
                            "cpf": social.responsible_cpf if social else None,
                            "telefone": social.phone if social else None,
                        },
                    }
                ],
            }

            zip_file.writestr(
                f"{folder_name}/geometria_{folder_name}.geojson",
                json.dumps(lot_geojson, ensure_ascii=False, indent=2),
            )

            zip_file.writestr(
                f"{folder_name}/geometria_{folder_name}.kml",
                _feature_collection_to_kml(
                    lot_geojson,
                    name=f"Geometria {lot.code}",
                ),
            )

            zip_file.writestr(
                f"{folder_name}/shp/geometria_{folder_name}.zip",
                _feature_collection_to_shapefile_zip(
                    lot_geojson,
                    name=f"geometria_{folder_name}"[:40],
                    geometry_type="polygon",
                ),
            )

        atributos_buffer = StringIO()
        atributos_writer = csv.writer(atributos_buffer, delimiter=";")

        atributos_writer.writerow(["campo", "valor"])
        atributos_writer.writerow(["codigo_lote", lot.code])
        atributos_writer.writerow(["quadra", lot.block])
        atributos_writer.writerow(["area_m2", lot.area_m2])
        atributos_writer.writerow(["perimetro_m", lot.perimeter_m])
        atributos_writer.writerow(["status_lote", lot.status])
        atributos_writer.writerow(
            ["status_tecnico", getattr(lot, "technical_status", None)]
        )
        atributos_writer.writerow(
            [
                "apto_pecas",
                _yes_no(getattr(lot, "is_ready_for_technical_documents", False)),
            ]
        )
        atributos_writer.writerow(["codigo_selo", seal.seal_code if seal else ""])
        atributos_writer.writerow(["situacao_selagem", seal.situation if seal else ""])
        atributos_writer.writerow(
            ["vinculo_geografico", seal.geo_link_status if seal else ""]
        )
        atributos_writer.writerow(
            ["responsavel", social.responsible_name if social else ""]
        )
        atributos_writer.writerow(["cpf", social.responsible_cpf if social else ""])
        atributos_writer.writerow(["rg", social.responsible_rg if social else ""])
        atributos_writer.writerow(["telefone", social.phone if social else ""])
        atributos_writer.writerow(
            ["tipo_imovel", physical.property_type if physical else ""]
        )
        atributos_writer.writerow(
            ["uso_imovel", physical.property_use if physical else ""]
        )
        atributos_writer.writerow(["documentos", documents_count])
        atributos_writer.writerow(["documentos_validados", validated_documents_count])

        zip_file.writestr(
            f"{folder_name}/atributos_{folder_name}.csv",
            atributos_buffer.getvalue(),
        )

        metadados = f"""DOSSIÊ INDIVIDUAL DO LOTE - BIOME REURB

Projeto: {project.name}
Município/UF: {project.municipality}/{project.state}
Bairro/Núcleo: {project.neighborhood}
Lote: {lot.code}
Data de geração: {datetime.now().strftime("%d/%m/%Y %H:%M")}

Conteúdo:
- Ficha cadastral PDF;
- Geometria individual em GeoJSON/KML/SHP, quando disponível;
- Atributos consolidados em CSV;
- Documentos vinculados ao lote/selagem/cadastro social, quando localizados no storage.

Observação:
Este dossiê subsidia a análise técnico-cadastral e documental do lote.
Não substitui planta topográfica, memorial descritivo ou ato administrativo final de regularização fundiária.
"""

        zip_file.writestr(
            f"{folder_name}/metadados_{folder_name}.txt",
            metadados,
        )

        for document in documents:
            document_path = _resolve_export_document_path(document)

            if document_path is None:
                continue

            original_name = (
                document.original_filename
                or document.stored_filename
                or document_path.name
            )

            safe_doc_name = _safe_filename_slug(Path(original_name).stem)
            suffix = Path(original_name).suffix or document_path.suffix

            archive_name = (
                f"{folder_name}/documentos/"
                f"{safe_doc_name}_{str(document.id)[:8]}{suffix}"
            )

            zip_file.write(document_path, arcname=archive_name)

    return FileResponse(
        path=zip_path,
        filename=filename,
        media_type="application/zip",
    )


@router.post("/orthomosaic/upload")
def upload_project_orthomosaic(
    project_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    project = db.query(Project).filter(Project.id == project_id).first()

    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projeto não encontrado.",
        )

    original_filename = file.filename or "ortomosaico.tif"
    suffix = Path(original_filename).suffix.lower()

    if suffix not in {".tif", ".tiff", ".geotiff"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Envie um arquivo GeoTIFF válido: .tif, .tiff ou .geotiff.",
        )

    ORTHOMOSAICS_DIR.mkdir(parents=True, exist_ok=True)

    project_dir = ORTHOMOSAICS_DIR / str(project_id)
    project_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stored_filename = f"ortomosaico_{timestamp}{suffix}"

    source_path = project_dir / stored_filename
    preview_path = project_dir / f"preview_{timestamp}.jpg"

    try:
        with source_path.open("wb") as output:
            while True:
                chunk = file.file.read(1024 * 1024)

                if not chunk:
                    break

                output.write(chunk)

        metadata = _generate_orthomosaic_preview_with_gdal(
            source_path=source_path,
            preview_path=preview_path,
        )

        db.query(ProjectOrthomosaic).filter(
            ProjectOrthomosaic.project_id == project_id,
            ProjectOrthomosaic.is_active == True,
        ).update({"is_active": False})

        orthomosaic = ProjectOrthomosaic(
            project_id=project_id,
            original_filename=original_filename,
            stored_filename=stored_filename,
            file_path=str(source_path),
            preview_path=str(preview_path),
            crs=metadata["crs"],
            min_lon=metadata["min_lon"],
            min_lat=metadata["min_lat"],
            max_lon=metadata["max_lon"],
            max_lat=metadata["max_lat"],
            width=metadata["width"],
            height=metadata["height"],
            is_active=True,
        )

        db.add(orthomosaic)
        db.commit()
        db.refresh(orthomosaic)

        return {
            "message": "Ortomosaico importado com sucesso.",
            "orthomosaic": _orthomosaic_to_response(orthomosaic),
        }

    except HTTPException:
        db.rollback()

        if source_path.exists():
            source_path.unlink(missing_ok=True)

        if preview_path.exists():
            preview_path.unlink(missing_ok=True)

        aux_source_path = Path(str(source_path) + ".aux.xml")
        aux_preview_path = Path(str(preview_path) + ".aux.xml")

        if aux_source_path.exists():
            aux_source_path.unlink(missing_ok=True)

        if aux_preview_path.exists():
            aux_preview_path.unlink(missing_ok=True)

        raise

    except Exception as exc:
        db.rollback()

        if source_path.exists():
            source_path.unlink(missing_ok=True)

        if preview_path.exists():
            preview_path.unlink(missing_ok=True)

        aux_source_path = Path(str(source_path) + ".aux.xml")
        aux_preview_path = Path(str(preview_path) + ".aux.xml")

        if aux_source_path.exists():
            aux_source_path.unlink(missing_ok=True)

        if aux_preview_path.exists():
            aux_preview_path.unlink(missing_ok=True)

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=(
                "Erro ao processar ortomosaico. Verifique se o arquivo é um GeoTIFF "
                "georreferenciado válido, com CRS reconhecido pelo GDAL. "
                f"Detalhe técnico: {exc}"
            ),
        )


@router.get("/orthomosaic")
def get_project_active_orthomosaic(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict | None:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    orthomosaic = (
        db.query(ProjectOrthomosaic)
        .filter(
            ProjectOrthomosaic.project_id == project_id,
            ProjectOrthomosaic.is_active == True,
        )
        .order_by(ProjectOrthomosaic.created_at.desc())
        .first()
    )

    if orthomosaic is None:
        return None

    return _orthomosaic_to_response(orthomosaic)


@router.get("/orthomosaic/{orthomosaic_id}/preview.png")
def get_project_orthomosaic_preview(
    project_id: UUID,
    orthomosaic_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> FileResponse:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    orthomosaic = (
        db.query(ProjectOrthomosaic)
        .filter(
            ProjectOrthomosaic.id == orthomosaic_id,
            ProjectOrthomosaic.project_id == project_id,
        )
        .first()
    )

    if orthomosaic is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ortomosaico não encontrado.",
        )

    preview_path = Path(orthomosaic.preview_path)

    if not preview_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Prévia do ortomosaico não encontrada.",
        )

    media_type = "image/jpeg"

    if preview_path.suffix.lower() == ".png":
        media_type = "image/png"

    return FileResponse(
        path=preview_path,
        filename=preview_path.name,
        media_type=media_type,
    )


@router.get("/orthomosaics")
def list_project_orthomosaics(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[dict]:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    items = (
        db.query(ProjectOrthomosaic)
        .filter(ProjectOrthomosaic.project_id == project_id)
        .order_by(
            ProjectOrthomosaic.is_active.desc(),
            ProjectOrthomosaic.created_at.desc(),
        )
        .all()
    )

    return [_orthomosaic_to_response(item) for item in items]


@router.patch("/orthomosaics/{orthomosaic_id}/activate")
def activate_project_orthomosaic(
    project_id: UUID,
    orthomosaic_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    orthomosaic = (
        db.query(ProjectOrthomosaic)
        .filter(
            ProjectOrthomosaic.id == orthomosaic_id,
            ProjectOrthomosaic.project_id == project_id,
        )
        .first()
    )

    if orthomosaic is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ortomosaico não encontrado.",
        )

    db.query(ProjectOrthomosaic).filter(
        ProjectOrthomosaic.project_id == project_id,
        ProjectOrthomosaic.is_active == True,
    ).update({"is_active": False})

    orthomosaic.is_active = True

    db.commit()
    db.refresh(orthomosaic)

    return {
        "message": "Ortomosaico definido como ativo.",
        "orthomosaic": _orthomosaic_to_response(orthomosaic),
    }


@router.delete("/orthomosaics/{orthomosaic_id}")
def delete_project_orthomosaic(
    project_id: UUID,
    orthomosaic_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _ensure_project_access(db, project_id=project_id, current_user=current_user)

    orthomosaic = (
        db.query(ProjectOrthomosaic)
        .filter(
            ProjectOrthomosaic.id == orthomosaic_id,
            ProjectOrthomosaic.project_id == project_id,
        )
        .first()
    )

    if orthomosaic is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ortomosaico não encontrado.",
        )

    was_active = bool(orthomosaic.is_active)

    paths_to_delete = [
        Path(orthomosaic.file_path),
        Path(orthomosaic.preview_path),
        Path(str(orthomosaic.file_path) + ".aux.xml"),
        Path(str(orthomosaic.preview_path) + ".aux.xml"),
    ]

    db.delete(orthomosaic)
    db.commit()

    for path in paths_to_delete:
        try:
            if path.exists():
                path.unlink(missing_ok=True)
        except Exception:
            # Não bloqueia a exclusão do registro se algum arquivo físico falhar.
            pass

    new_active = None

    if was_active:
        new_active = (
            db.query(ProjectOrthomosaic)
            .filter(ProjectOrthomosaic.project_id == project_id)
            .order_by(ProjectOrthomosaic.created_at.desc())
            .first()
        )

        if new_active is not None:
            new_active.is_active = True
            db.commit()
            db.refresh(new_active)

    return {
        "message": "Ortomosaico excluído com sucesso.",
        "new_active_orthomosaic": (
            _orthomosaic_to_response(new_active) if new_active else None
        ),
    }


@router.get(
    "/seals/{seal_id}/delete-check",
    response_model=SealDeleteCheckResponse,
)
def check_seal_deletion(
    project_id: UUID,
    seal_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> SealDeleteCheckResponse:
    _ensure_administrative_access(current_user)

    seal = (
        db.query(Seal)
        .filter(
            Seal.id == seal_id,
            Seal.project_id == project_id,
            Seal.deleted.is_(False),
        )
        .first()
    )

    if seal is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Selagem não encontrada ou já excluída.",
        )

    social_count = (
        db.query(func.count(SocialRegistration.id))
        .filter(
            SocialRegistration.project_id == project_id,
            SocialRegistration.seal_id == seal.id,
        )
        .scalar()
        or 0
    )

    physical_count = (
        db.query(func.count(PhysicalRegistration.id))
        .filter(
            PhysicalRegistration.project_id == project_id,
            PhysicalRegistration.seal_id == seal.id,
        )
        .scalar()
        or 0
    )

    document_count = (
        db.query(func.count(Document.id))
        .filter(
            Document.project_id == project_id,
            Document.seal_id == seal.id,
        )
        .scalar()
        or 0
    )

    geometry_count = (
        db.query(func.count(LotGeometry.id))
        .filter(
            LotGeometry.project_id == project_id,
            LotGeometry.seal_id == seal.id,
            LotGeometry.deleted.is_(False),
        )
        .scalar()
        or 0
    )

    return SealDeleteCheckResponse(
        seal_id=seal.id,
        seal_code=seal.seal_code,
        social_registrations=int(social_count),
        physical_registrations=int(physical_count),
        documents=int(document_count),
        lot_geometries=int(geometry_count),
        linked_lot=seal.lot_id is not None,
    )


@router.delete(
    "/seals/{seal_id}",
    response_model=SealDeleteResponse,
)
def delete_seal_administratively(
    project_id: UUID,
    seal_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> SealDeleteResponse:
    _ensure_administrative_access(current_user)

    seal = (
        db.query(Seal)
        .filter(
            Seal.id == seal_id,
            Seal.project_id == project_id,
        )
        .with_for_update()
        .first()
    )

    if seal is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Selagem não encontrada.",
        )

    if seal.deleted:
        return SealDeleteResponse(
            seal_id=seal.id,
            seal_code=seal.seal_code,
            message="A selagem já estava excluída.",
        )

    try:
        social_count = (
            db.query(func.count(SocialRegistration.id))
            .filter(
                SocialRegistration.project_id == project_id,
                SocialRegistration.seal_id == seal.id,
            )
            .scalar()
            or 0
        )

        physical_count = (
            db.query(func.count(PhysicalRegistration.id))
            .filter(
                PhysicalRegistration.project_id == project_id,
                PhysicalRegistration.seal_id == seal.id,
            )
            .scalar()
            or 0
        )

        document_count = (
            db.query(func.count(Document.id))
            .filter(
                Document.project_id == project_id,
                Document.seal_id == seal.id,
            )
            .scalar()
            or 0
        )

        geometries = (
            db.query(LotGeometry)
            .filter(
                LotGeometry.project_id == project_id,
                LotGeometry.seal_id == seal.id,
                LotGeometry.deleted.is_(False),
            )
            .all()
        )

        now = datetime.now(timezone.utc)

        for geometry in geometries:
            geometry.deleted = True
            geometry.is_current = False
            geometry.workflow_status = "arquivado"
            geometry.validation_note = (
                "Geometria arquivada automaticamente porque a "
                "selagem vinculada foi excluída pelo painel web."
            )
            geometry.updated_at = now

        db.query(Document).filter(
            Document.project_id == project_id,
            Document.seal_id == seal.id,
        ).delete(synchronize_session=False)

        db.query(PhysicalRegistration).filter(
            PhysicalRegistration.project_id == project_id,
            PhysicalRegistration.seal_id == seal.id,
        ).delete(synchronize_session=False)

        db.query(SocialRegistration).filter(
            SocialRegistration.project_id == project_id,
            SocialRegistration.seal_id == seal.id,
        ).delete(synchronize_session=False)

        seal.lot_id = None
        seal.deleted = True
        seal.sync_version = (seal.sync_version or 0) + 1
        seal.updated_by_user_id = current_user.id
        seal.server_received_at = now
        seal.updated_at = now

        db.commit()

        register_audit_log(
            db,
            user_id=current_user.id,
            action="delete",
            entity_type="seal",
            entity_id=seal.id,
            project_id=project_id,
            details={
                "seal_code": seal.seal_code,
                "social_registrations_removed": int(social_count),
                "physical_registrations_removed": int(physical_count),
                "documents_removed": int(document_count),
                "lot_geometries_archived": len(geometries),
            },
        )

        return SealDeleteResponse(
            seal_id=seal.id,
            seal_code=seal.seal_code,
            social_registrations_removed=int(social_count),
            physical_registrations_removed=int(physical_count),
            documents_removed=int(document_count),
            lot_geometries_archived=len(geometries),
            message=(
                "Selagem excluída e dependências tratadas. "
                "A exclusão será refletida nos aplicativos na próxima sincronização."
            ),
        )

    except Exception as exc:
        db.rollback()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Não foi possível excluir a selagem: {exc}",
        ) from exc
